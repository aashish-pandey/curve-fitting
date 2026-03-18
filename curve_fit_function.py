"""
curve_fit_function.py
─────────────────────
Drop-in module for NMR curve fitting over multiple Excel sheets.

Usage
-----
    from curve_fit_function import fit_curves

    out_path = fit_curves("my_data.xlsx")                         # defaults
    out_path = fit_curves("my_data.xlsx", x_col="Bvalue",
                          y_col="integrals",
                          output_path="my_data_fitted.xlsx")

Parameters
----------
excel_path  : str  – path to the input .xlsx workbook
x_col       : str  – header of the x-axis column  (default "Bvalue")
y_col       : str  – header of the y-axis column  (default "integrals")
output_path : str  – where to save the result
                     (default: <input_stem>_fitted.xlsx next to the input)

Returns
-------
str – absolute path of the saved output workbook

For every sheet the function
  • locates x / y columns by header name  (case-insensitive, strips whitespace)
  • fits 8 NMR models via Nelder-Mead or OLS
  • appends M1_fit … M8_fit  and  M1_res … M8_res  columns
  • writes a model-summary table below the data in the same sheet

Models
------
  1  Mono-Exp          y = B · exp(−F·x)
  2  Mono-Exp+Offset   y = B + F · exp(−G·x)
  3  Bi-Exp            y = B1·exp(−F1·x) + B2·exp(−F2·x)
  4  Linear            y = A + B·x                       (OLS)
  5  Inv.Linear        y = 1 / (A + B·x)
  6  Inv.Lin+Offset    y = 1 / (A + B·x) + C
  7  Intermediate      y = I·(exp(−K·x)−exp(−Q·x)) / (Q/K−1)
  8  Interm+Offset     y = model_7 + C

Dependencies: math, os  (stdlib only) + openpyxl
"""

import math
import os

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════════════
def fit_curves(
    excel_path: str,
    x_col: str = "Bvalue",
    y_col: str = "integrals",
    output_path: str = None,
) -> str:
    """Fit 8 NMR curves to every sheet; return path of saved workbook."""

    # ─────────────────────────────────────────────────────────────────────────
    # SECTION 1 – PURE-MATH HELPERS
    # ─────────────────────────────────────────────────────────────────────────

    def _clip(v, lo=-700.0, hi=700.0):
        """Clamp v to [lo, hi] so math.exp never overflows."""
        return lo if v < lo else (hi if v > hi else v)

    def _r2(y_true, y_pred):
        """Coefficient of determination R²."""
        ym = sum(y_true) / len(y_true)
        ss_tot = sum((y - ym) ** 2 for y in y_true)
        if ss_tot == 0:
            return 0.0
        ss_res = sum(
            (yt - yp) ** 2
            for yt, yp in zip(y_true, y_pred)
            if yp is not None and not math.isnan(yp)
        )
        return 1.0 - ss_res / ss_tot

    def _ssr(y_true, y_pred):
        """Sum of squared residuals (skips None / NaN predictions)."""
        return sum(
            (yt - yp) ** 2
            for yt, yp in zip(y_true, y_pred)
            if yp is not None and not math.isnan(yp)
        )

    def _fp(v):
        """Adaptive number formatter: sci-notation for very small/large values."""
        if not isinstance(v, (int, float)):
            return str(v)
        if math.isinf(v) or math.isnan(v):
            return str(v)
        if v == 0.0:
            return "0"
        if abs(v) < 1e-3 or abs(v) > 9999:
            return f"{v:.4e}"
        return f"{v:.6f}"

    # ─────────────────────────────────────────────────────────────────────────
    # SECTION 2 – NELDER-MEAD OPTIMISER  (pure Python)
    # ─────────────────────────────────────────────────────────────────────────

    def _nelder_mead(f, x0, tol=1e-12, max_iter=20000):
        """Minimise scalar function f starting at vector x0."""
        n = len(x0)

        # Build initial simplex – each vertex perturbs one parameter
        simplex = [list(x0)]
        for i in range(n):
            pt = list(x0)
            pt[i] += 0.1 * abs(pt[i]) if abs(pt[i]) > 1e-10 else 1e-4
            simplex.append(pt)
        scores = [f(s) for s in simplex]

        alpha, gamma, rho, sigma = 1.0, 2.0, 0.5, 0.5

        for _ in range(max_iter):
            order   = sorted(range(n + 1), key=lambda i: scores[i])
            simplex = [simplex[i] for i in order]
            scores  = [scores[i]  for i in order]

            if scores[-1] - scores[0] < tol:
                break

            # centroid of all but worst vertex
            c = [sum(simplex[i][j] for i in range(n)) / n for j in range(n)]

            # reflection
            xr = [c[j] + alpha * (c[j] - simplex[-1][j]) for j in range(n)]
            fr = f(xr)

            if fr < scores[0]:                        # expansion
                xe = [c[j] + gamma * (xr[j] - c[j]) for j in range(n)]
                fe = f(xe)
                simplex[-1], scores[-1] = (xe, fe) if fe < fr else (xr, fr)
            elif fr < scores[-2]:                     # accept reflection
                simplex[-1], scores[-1] = xr, fr
            else:                                     # contraction
                if fr < scores[-1]:
                    simplex[-1], scores[-1] = xr, fr
                xc = [c[j] + rho * (simplex[-1][j] - c[j]) for j in range(n)]
                fc = f(xc)
                if fc < scores[-1]:
                    simplex[-1], scores[-1] = xc, fc
                else:                                 # shrink
                    for i in range(1, n + 1):
                        simplex[i] = [
                            simplex[0][j] + sigma * (simplex[i][j] - simplex[0][j])
                            for j in range(n)
                        ]
                        scores[i] = f(simplex[i])

        return simplex[0]

    # ─────────────────────────────────────────────────────────────────────────
    # SECTION 3 – GAUSSIAN ELIMINATION + OLS  (pure Python)
    # ─────────────────────────────────────────────────────────────────────────

    def _gauss_solve(A, b):
        """Solve Ax = b via Gaussian elimination with partial pivoting."""
        n = len(b)
        M = [A[i][:] + [b[i]] for i in range(n)]
        for col in range(n):
            pivot = max(range(col, n), key=lambda r: abs(M[r][col]))
            M[col], M[pivot] = M[pivot], M[col]
            if abs(M[col][col]) < 1e-15:
                continue
            for row in range(col + 1, n):
                f_ = M[row][col] / M[col][col]
                M[row] = [M[row][k] - f_ * M[col][k] for k in range(n + 1)]
        x = [0.0] * n
        for i in range(n - 1, -1, -1):
            x[i] = (
                M[i][n] - sum(M[i][j] * x[j] for j in range(i + 1, n))
            ) / M[i][i]
        return x

    def _ols(xs, ys, features_fn):
        """Ordinary least squares.  features_fn(x) -> list of feature values."""
        rows = [features_fn(xi) for xi in xs]
        k    = len(rows[0])
        XtX  = [[sum(r[a] * r[b] for r in rows) for b in range(k)] for a in range(k)]
        Xty  = [sum(rows[i][j] * ys[i] for i in range(len(ys))) for j in range(k)]
        return _gauss_solve(XtX, Xty)

    # ─────────────────────────────────────────────────────────────────────────
    # SECTION 4 – MODEL FUNCTIONS  (pure-Python scalars)
    # ─────────────────────────────────────────────────────────────────────────

    def m1(x, B, F):
        return B * math.exp(_clip(-x * F))

    def m2(x, B, F, G):
        return B + F * math.exp(_clip(-x * G))

    def m3(x, B1, F1, B2, F2):
        return (B1 * math.exp(_clip(-x * F1))
                + B2 * math.exp(_clip(-x * F2)))

    def m4(x, A, B):           # linear — fitted via OLS
        return A + B * x

    def m5(x, A, B):
        d = A + B * x
        return 1.0 / (d if d != 0 else 1e-12)

    def m6(x, A, B, C):
        d = A + B * x
        return 1.0 / (d if d != 0 else 1e-12) + C

    def m7(x, I, K, Q):
        eK    = math.exp(_clip(-x * K))
        eQ    = math.exp(_clip(-x * Q))
        K_    = K if K != 0 else 1e-12
        denom = (Q / K_) - 1.0
        if abs(denom) < 1e-8:           # L'Hôpital limit when Q ≈ K
            return I * x * math.exp(_clip(-K * x))
        return I * (eK - eQ) / denom

    def m8(x, I, K, Q, C):
        return m7(x, I, K, Q) + C

    # Metadata kept in one place so loops stay DRY
    # col  = prefix used for Excel column headers  (no spaces / special chars)
    _MODEL_META = {
        1: dict(name="Mono-Exp",           col="MonoExp",
                formula="y = B·exp(−F·x)",
                pnames=["B","F"]),
        2: dict(name="Mono-Exp+Offset",    col="MonoExpOffset",
                formula="y = B + F·exp(−G·x)",
                pnames=["B","F","G"]),
        3: dict(name="Bi-Exp",             col="BiExp",
                formula="y = B1·exp(−F1·x) + B2·exp(−F2·x)",
                pnames=["B1","F1","B2","F2"]),
        4: dict(name="Linear",             col="Linear",
                formula="y = A + B·x  (OLS)",
                pnames=["A","B"]),
        5: dict(name="Inv.Linear",         col="InvLinear",
                formula="y = 1/(A + B·x)",
                pnames=["A","B"]),
        6: dict(name="Inv.Lin+Offset",     col="InvLinearOffset",
                formula="y = 1/(A + B·x) + C",
                pnames=["A","B","C"]),
        7: dict(name="Intermediate",       col="Intermediate",
                formula="y = I·(exp(−K·x)−exp(−Q·x))/(Q/K−1)",
                pnames=["I","K","Q"]),
        8: dict(name="Interm+Offset",      col="IntermediateOffset",
                formula="y = model_7 + C",
                pnames=["I","K","Q","C"]),
    }

    # ─────────────────────────────────────────────────────────────────────────
    # SECTION 5 – PER-SHEET PROCESSING
    # ─────────────────────────────────────────────────────────────────────────

    def _find_columns(ws, x_name, y_name):
        """
        Scan the first 10 rows to find x and y column indices (1-based).
        Returns (header_row, x_col_idx, y_col_idx) or raises ValueError.
        """
        x_low = x_name.strip().lower()
        y_low = y_name.strip().lower()
        for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row)):
            found = {}
            for cell in row:
                if cell.value is None:
                    continue
                label = str(cell.value).strip().lower()
                if label == x_low:
                    found["x"] = cell.column
                elif label == y_low:
                    found["y"] = cell.column
            if "x" in found and "y" in found:
                return row[0].row, found["x"], found["y"]
        raise ValueError(
            f"Columns '{x_name}' and '{y_name}' not found in sheet '{ws.title}'"
        )

    def _read_data(ws, header_row, x_col_idx, y_col_idx):
        """
        Read numeric x/y pairs from the sheet.
        Returns (x_data, y_data, row_numbers).
        row_numbers[i] = actual worksheet row index for data point i.
        """
        x_data, y_data, row_nums = [], [], []
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
            xv = row[x_col_idx - 1].value
            yv = row[y_col_idx - 1].value
            if xv is None or yv is None:
                continue
            try:
                x_data.append(float(xv))
                y_data.append(float(yv))
                row_nums.append(row[0].row)
            except (ValueError, TypeError):
                continue
        return x_data, y_data, row_nums

    def _warm_starts(x_data, y_data):
        """
        Derive data-driven warm-start estimates for Nelder-Mead.
        Returns dict with keys: amp, rate, inv_A, inv_B
        """
        srt     = sorted(zip(x_data, y_data))
        xf, yf  = srt[0]
        xl, yl  = srt[-1]
        y_pos   = [y for y in y_data if y > 0]

        amp = yf if yf > 0 else (sum(y_pos) / len(y_pos) if y_pos else 1.0)
        dx  = abs(xl - xf) if abs(xl - xf) > 1e-12 else 1.0
        yf_ = yf if abs(yf) > 1e-30 else 1e-30
        yl_ = yl if abs(yl) > 1e-30 else 1e-30
        rate = abs(math.log(abs(yl_ / yf_))) / dx
        if rate < 1e-30:
            rate = 1e-5

        inv_A = 1.0 / amp if amp != 0 else 1.0
        inv_B = (1.0 / yl_ - inv_A) / dx if dx > 0 else 0.0

        return dict(amp=amp, rate=rate, inv_A=inv_A, inv_B=inv_B)

    def _fit_all(x_data, y_data):
        """
        Fit all 8 models. Returns dict:
            model_num -> (params, y_pred, r2, residuals)
        """

        def fit_nm(model_fn, p0):
            """Nelder-Mead fit; returns (params, y_pred, r2, residuals)."""
            def ssr_fn(p):
                try:
                    return sum(
                        (yi - model_fn(xi, *p)) ** 2
                        for xi, yi in zip(x_data, y_data)
                    )
                except (OverflowError, ValueError, ZeroDivisionError):
                    return 1e30

            params = _nelder_mead(ssr_fn, p0)

            y_pred = []
            for xi in x_data:
                try:
                    v = model_fn(xi, *params)
                    y_pred.append(v if not math.isnan(v) else None)
                except (OverflowError, ValueError, ZeroDivisionError):
                    y_pred.append(None)

            resid = [
                (yo - yp) if yp is not None else None
                for yo, yp in zip(y_data, y_pred)
            ]
            return params, y_pred, _r2(y_data, y_pred), resid

        def fit_ols_linear():
            """Exact OLS for the linear model."""
            coeff = _ols(x_data, y_data, lambda x: [1.0, x])
            A, B  = coeff[0], coeff[1]
            yp    = [m4(xi, A, B) for xi in x_data]
            resid = [yo - yp_v for yo, yp_v in zip(y_data, yp)]
            return (A, B), yp, _r2(y_data, yp), resid

        ws_ = _warm_starts(x_data, y_data)
        amp, rate, inv_A, inv_B = ws_["amp"], ws_["rate"], ws_["inv_A"], ws_["inv_B"]

        out = {}
        out[1] = fit_nm(m1, [amp, rate])
        out[2] = fit_nm(m2, [0.0, amp, rate])
        out[3] = fit_nm(m3, [amp * 0.7, rate, amp * 0.3, rate * 3.0])
        out[4] = fit_ols_linear()
        out[5] = fit_nm(m5, [inv_A, inv_B])
        out[6] = fit_nm(m6, [inv_A, inv_B, 0.0])
        out[7] = fit_nm(m7, [amp, rate, rate * 10.0])
        out[8] = fit_nm(m8, [amp, rate, rate * 10.0, 0.0])
        return out

    def _write_fitted_columns(ws, header_row, y_col_idx, row_nums, results):
        """
        Append M1_fit / M1_res … M8_fit / M8_res columns to the right of
        existing data.  Residuals are stored as live Excel formulas.

        Returns dict:  model_num -> (fit_col_letter, res_col_letter)
        """
        start_col    = ws.max_column + 1
        y_col_letter = get_column_letter(y_col_idx)
        col_map      = {}

        for i, num in enumerate(range(1, 9)):
            fc = start_col + i * 2          # fitted column index
            rc = fc + 1                     # residual column index
            fc_letter = get_column_letter(fc)
            rc_letter = get_column_letter(rc)
            col_map[num] = (fc_letter, rc_letter)

            # headers — use descriptive name from metadata
            col_prefix = _MODEL_META[num]["col"]
            hdr_fit = ws.cell(row=header_row, column=fc, value=f"{col_prefix}_fit")
            hdr_res = ws.cell(row=header_row, column=rc, value=f"{col_prefix}_res")
            hdr_fit.font = Font(bold=True)
            hdr_res.font = Font(bold=True)

            # data rows
            params, y_pred, r2, resid = results[num]
            for row_num, yp_val in zip(row_nums, y_pred):
                ws.cell(row=row_num, column=fc,
                        value=round(yp_val, 12) if yp_val is not None else None)
                # live formula so the residual updates if the user edits data
                ws.cell(row=row_num, column=rc,
                        value=f"={y_col_letter}{row_num}-{fc_letter}{row_num}")

        return col_map

    def _write_summary(ws, last_data_row, x_data, y_data, results):
        """
        Write a compact model-summary table starting 3 rows below the data.
        Columns: #  |  Name  |  Formula  |  R²  |  SSR  |  Parameters
        """
        # ── style constants ────────────────────────────────────────────────
        FILL_HDR   = PatternFill(fill_type="solid", fgColor="2F5496")
        FILL_ALT   = PatternFill(fill_type="solid", fgColor="D9E1F2")
        FONT_WHITE = Font(bold=True, color="FFFFFF")
        FONT_BOLD  = Font(bold=True)
        CENTRE     = Alignment(horizontal="center")

        start_row = last_data_row + 3
        headers   = ["#", "Name", "Formula", "R²", "SSR", "Parameters"]

        # header row
        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=ci, value=h)
            cell.font      = FONT_WHITE
            cell.fill      = FILL_HDR
            cell.alignment = CENTRE

        # one row per model
        for num in range(1, 9):
            r      = start_row + num
            params, y_pred, r2, _ = results[num]
            ssr    = _ssr(y_data, y_pred)
            meta   = _MODEL_META[num]
            pnames = meta["pnames"]
            pstr   = ",  ".join(f"{n}={_fp(v)}" for n, v in zip(pnames, params))

            row_vals = [num, meta["name"], meta["formula"],
                        round(r2, 8), ssr, pstr]
            for ci, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=r, column=ci, value=val)
                if num % 2 == 0:
                    cell.fill = FILL_ALT
                if ci == 1:
                    cell.alignment = CENTRE

        # reasonable column widths for the summary block
        for col, width in zip(range(1, 7), [5, 18, 44, 14, 14, 48]):
            ws.column_dimensions[get_column_letter(col)].width = width

    def process_sheet(ws):
        """Full pipeline for a single worksheet."""
        # -- locate columns --------------------------------------------------
        try:
            header_row, x_idx, y_idx = _find_columns(ws, x_col, y_col)
        except ValueError as e:
            print(f"  [skip] {e}")
            return

        # -- read data -------------------------------------------------------
        x_data, y_data, row_nums = _read_data(ws, header_row, x_idx, y_idx)
        if len(x_data) < 4:
            print(f"  [skip] sheet '{ws.title}': only {len(x_data)} valid rows (need ≥4).")
            return
        print(f"  Fitting sheet '{ws.title}': {len(x_data)} points …")

        # -- fit all 8 models ------------------------------------------------
        results = _fit_all(x_data, y_data)

        # -- print console summary -------------------------------------------
        for num in range(1, 9):
            params, _, r2, _ = results[num]
            meta   = _MODEL_META[num]
            pnames = meta["pnames"]
            pstr   = ", ".join(f"{n}={_fp(v)}" for n, v in zip(pnames, params))
            print(f"    M{num}  {meta['name']:<18} R²={r2:.6f}   {pstr}")

        # -- write to sheet --------------------------------------------------
        _write_fitted_columns(ws, header_row, y_idx, row_nums, results)
        last_row = max(row_nums)
        _write_summary(ws, last_row, x_data, y_data, results)

    # ─────────────────────────────────────────────────────────────────────────
    # SECTION 6 – MAIN EXECUTION
    # ─────────────────────────────────────────────────────────────────────────

    wb = openpyxl.load_workbook(excel_path)
    print(f"Workbook: {excel_path}  ({len(wb.sheetnames)} sheet(s))")

    for sheet_name in wb.sheetnames:
        process_sheet(wb[sheet_name])

    if output_path is None:
        base, ext = os.path.splitext(os.path.abspath(excel_path))
        output_path = base + "_fitted" + (ext if ext else ".xlsx")

    wb.save(output_path)
    print(f"\nSaved → {output_path}")
    return output_path


# ──────────────────────────────────────────────────────────────────────────────
# Allow running directly for quick testing:
#   python curve_fit_function.py data.xlsx Bvalue integrals
# ──────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys

    _path  = sys.argv[1] if len(sys.argv) > 1 else "monoexp_test_data.xlsx"
    _xcol  = sys.argv[2] if len(sys.argv) > 2 else "integral"
    _ycol  = sys.argv[3] if len(sys.argv) > 3 else "b_value"
    _out   = sys.argv[4] if len(sys.argv) > 4 else None

    fit_curves(_path, x_col=_xcol, y_col=_ycol, output_path=_out)
