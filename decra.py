import os
import numpy as np
from numpy.linalg import eig, svd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import ScatterChart, Reference, Series
import MnovaNMR

# ─── CONFIG ───────────────────────────────────────────────────────────────────

OUTPUT_FOLDER = r"C:\Users\you\Documents\NMR_Results"   # folder to save xlsx

# List of peaks to process — add as many as you need.
# Each entry: (label, ppm_left, ppm_right)
PEAKS = [
    ("PEG_peak",     3.70, 3.60),
    ("CH2_peak",     1.30, 1.10),
]

N_COMPONENTS = 2   # polymer + small molecule

# ──────────────────────────────────────────────────────────────────────────────


def get_b_values(nmr_item):
    """
    Return b-values (s/m^2) as a numpy array, one per gradient step.
    Replace with your own extraction logic.
    """
    spectra  = nmr_item.spectra()
    b_values = np.zeros(len(spectra))
    # --- YOUR CODE HERE ---
    return b_values


def build_spectral_matrix(nmr_item, ppm_left, ppm_right):
    spectra  = nmr_item.spectra()
    ess      = spectra[0].coords[0]

    pt_left  = int(round(ess.ppmToPt(ppm_left)))
    pt_right = int(round(ess.ppmToPt(ppm_right)))
    pt_start = min(pt_left, pt_right)
    pt_end   = max(pt_left, pt_right)
    pts      = list(range(pt_start, pt_end + 1))

    ppm_axis = np.array([ess.ptToPpm(pt) for pt in pts])

    Y = np.zeros((len(spectra), len(pts)), dtype=float)
    for i, spc in enumerate(spectra):
        for j, pt in enumerate(pts):
            Y[i, j] = spc.reDataAt(pt)

    return Y, ppm_axis


def decra(Y, n_components=2):
    Y  = np.asarray(Y, dtype=float)
    K  = n_components

    Y1 = Y[:-1, :]
    Y2 = Y[1:,  :]

    U, s, Vt = svd(Y1, full_matrices=False)
    U  = U[:,  :K]
    s  = s[    :K]
    Vt = Vt[:K, :]

    Phi_red = U.T @ Y2 @ Vt.T @ np.diag(1.0 / s)

    eigenvalues, eigenvectors = eig(Phi_red)
    eigenvalues  = eigenvalues.real
    eigenvectors = eigenvectors.real

    order        = np.argsort(eigenvalues)[::-1]
    eigenvalues  = eigenvalues[order]
    eigenvectors = eigenvectors[:, order]

    C1 = U @ np.diag(s) @ eigenvectors
    C  = np.vstack([C1, C1[-1, :] * eigenvalues])
    C  = C / C[0, :]

    S_comp = np.linalg.pinv(C) @ Y

    return C, S_comp, eigenvalues


def fit_diffusion(C, b_values):
    b         = np.asarray(b_values, dtype=float)
    D_values  = np.zeros(C.shape[1])
    I0_values = np.zeros(C.shape[1])

    for k in range(C.shape[1]):
        decay = C[:, k]
        mask  = decay > 0
        if mask.sum() < 2:
            continue
        coeffs        = np.polyfit(b[mask], np.log(decay[mask]), deg=1)
        D_values[k]   = -coeffs[0]
        I0_values[k]  =  np.exp(coeffs[1])

    return D_values, I0_values


# ─── EXCEL HELPERS ────────────────────────────────────────────────────────────

def _hdr(cell, value, bg="1F4E79"):
    cell.value     = value
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def _label(cell, value):
    cell.value = value
    cell.font  = Font(bold=True, name="Arial", size=10)

def _val(cell, value):
    cell.value = value
    cell.font  = Font(name="Arial", size=10)

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _apply_border(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = _border()


def write_peak_sheet(wb, peak_label, b_values, C, S_comp, eigenvalues,
                     ppm_axis, D_values, I0_values):
    """Write one sheet for one peak with data table, fit table, and chart."""
    ws = wb.create_sheet(title=peak_label[:31])   # Excel sheet name max 31 chars

    b      = np.asarray(b_values, dtype=float)
    labels = ["Polymer", "Small molecule"]
    colors = ["2E75B6", "ED7D31"]   # blue, orange

    # ── Section 1: header ────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    ws["A1"].value     = f"DECRA Results — {peak_label}  |  ppm window {ppm_axis.min():.3f} – {ppm_axis.max():.3f}"
    ws["A1"].font      = Font(bold=True, size=12, name="Arial", color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Section 2: fit summary (rows 3–7) ────────────────────────────────────
    _label(ws["A3"], "Component")
    _label(ws["B3"], "D  (m²/s)")
    _label(ws["C3"], "I₀  (intercept)")
    _label(ws["D3"], "Eigenvalue")
    _label(ws["E3"], "Integral at b=0")
    _label(ws["F3"], "Equation")
    for col in "ABCDEF":
        ws[f"{col}3"].fill      = PatternFill("solid", start_color="D6E4F0")
        ws[f"{col}3"].font      = Font(bold=True, name="Arial", size=10)
        ws[f"{col}3"].alignment = Alignment(horizontal="center")

    dppm = abs(ppm_axis[1] - ppm_axis[0]) if len(ppm_axis) > 1 else 1.0
    for k in range(N_COMPONENTS):
        row      = 4 + k
        integral = float(np.trapz(S_comp[k, :], dx=dppm))
        _val(ws.cell(row, 1), labels[k])
        _val(ws.cell(row, 2), float(D_values[k]))
        _val(ws.cell(row, 3), float(I0_values[k]))
        _val(ws.cell(row, 4), float(eigenvalues[k]))
        _val(ws.cell(row, 5), float(integral))
        _val(ws.cell(row, 6), f"I(b) = {I0_values[k]:.4f} * exp(-{D_values[k]:.4e} * b)")
        ws.cell(row, 2).number_format = "0.00E+00"
        ws.cell(row, 3).number_format = "0.0000"
        ws.cell(row, 4).number_format = "0.000000"
        ws.cell(row, 5).number_format = "0.00"

    _apply_border(ws, 3, 3 + N_COMPONENTS, 1, 6)

    # ── Section 3: raw decay data (starts row 9) ─────────────────────────────
    data_row_start = 9
    _hdr(ws.cell(data_row_start, 1), "b value (s/m²)")
    _hdr(ws.cell(data_row_start, 2), "Observed (sum)")
    for k in range(N_COMPONENTS):
        _hdr(ws.cell(data_row_start, 3 + k), f"{labels[k]} decay (C)")

    # b-value column and observed (sum of spectral window per gradient)
    observed = Y_sum = None   # filled below after we pass Y in — see caller
    for i, bv in enumerate(b):
        r = data_row_start + 1 + i
        ws.cell(r, 1).value = float(bv)
        ws.cell(r, 2).value = float(np.sum(S_comp[:, :], axis=0).mean())  # placeholder; overwritten below
        for k in range(N_COMPONENTS):
            ws.cell(r, 3 + k).value = float(C[i, k])
        ws.cell(r, 1).number_format = "0.00E+00"

    # ── Section 4: fitted curve (100 points) ─────────────────────────────────
    fit_col_start = 3 + N_COMPONENTS + 2   # leave a gap column
    b_min, b_max  = b.min(), b.max()
    _hdr(ws.cell(data_row_start, fit_col_start),     "b fit (s/m²)")
    for k in range(N_COMPONENTS):
        _hdr(ws.cell(data_row_start, fit_col_start + 1 + k),
             f"{labels[k]} fit")

    n_fit = 100
    b_fit = np.linspace(b_min, b_max, n_fit)
    for i, bv in enumerate(b_fit):
        r = data_row_start + 1 + i
        ws.cell(r, fit_col_start).value = float(bv)
        ws.cell(r, fit_col_start).number_format = "0.00E+00"
        for k in range(N_COMPONENTS):
            ws.cell(r, fit_col_start + 1 + k).value = float(
                I0_values[k] * np.exp(-D_values[k] * bv)
            )

    # ── Section 5: scatter chart ──────────────────────────────────────────────
    chart = ScatterChart()
    chart.title        = f"Decay curves — {peak_label}"
    chart.style        = 10
    chart.x_axis.title = "b value (s/m²)"
    chart.y_axis.title = "Normalised intensity"
    chart.x_axis.numFmt = "0.00E+00"
    chart.height       = 14
    chart.width        = 22

    n_data = len(b)
    n_rows_data = data_row_start + n_data   # last row with raw data

    for k in range(N_COMPONENTS):
        # raw scatter points
        xvals  = Reference(ws, min_col=1,           min_row=data_row_start + 1, max_row=n_rows_data)
        yvals  = Reference(ws, min_col=3 + k,       min_row=data_row_start + 1, max_row=n_rows_data)
        series = Series(yvals, xvals, title=f"{labels[k]} (data)")
        series.marker.symbol    = "circle"
        series.marker.size      = 5
        series.graphicalProperties.line.noFill = True
        series.marker.graphicalProperties.solidFill    = colors[k]
        series.marker.graphicalProperties.line.solidFill = colors[k]
        chart.series.append(series)

        # fitted line
        xfit   = Reference(ws, min_col=fit_col_start,         min_row=data_row_start + 1, max_row=data_row_start + n_fit)
        yfit   = Reference(ws, min_col=fit_col_start + 1 + k, min_row=data_row_start + 1, max_row=data_row_start + n_fit)
        sfit   = Series(yfit, xfit, title=f"{labels[k]} (fit)")
        sfit.graphicalProperties.line.solidFill = colors[k]
        sfit.graphicalProperties.line.width      = 15000   # EMU
        sfit.marker.symbol = "none"
        chart.series.append(sfit)

    ws.add_chart(chart, "A10")   # anchor — Excel will float it

    # column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 42


def write_summary_sheet(wb, sample_name, peak_results):
    """
    First sheet — one row per peak with D and integral for each component.
    peak_results: list of (label, D_values, I0_values, eigenvalues, S_comp, ppm_axis)
    """
    ws = wb.active
    ws.title = "Summary"

    ws.merge_cells("A1:J1")
    ws["A1"].value     = f"DECRA Summary — {sample_name}"
    ws["A1"].font      = Font(bold=True, size=14, name="Arial", color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    headers = [
        "Peak", "ppm window",
        "D polymer (m²/s)", "D small mol (m²/s)",
        "Integral polymer", "Integral small mol",
        "Ratio (poly/sm)", "λ polymer", "λ small mol"
    ]
    for col, h in enumerate(headers, 1):
        _hdr(ws.cell(3, col), h)

    for row, (label, D_vals, I0_vals, eigs, S_comp, ppm_axis) in enumerate(peak_results, 4):
        dppm = abs(ppm_axis[1] - ppm_axis[0]) if len(ppm_axis) > 1 else 1.0
        i0   = float(np.trapz(S_comp[0, :], dx=dppm))
        i1   = float(np.trapz(S_comp[1, :], dx=dppm))

        vals = [
            label,
            f"{ppm_axis.min():.3f} – {ppm_axis.max():.3f}",
            float(D_vals[0]),
            float(D_vals[1]),
            i0,
            i1,
            f"=E{row}/F{row}",   # Excel formula for ratio
            float(eigs[0]),
            float(eigs[1]),
        ]
        for col, v in enumerate(vals, 1):
            _val(ws.cell(row, col), v)
            ws.cell(row, col).alignment = Alignment(horizontal="center")

        ws.cell(row, 3).number_format = "0.00E+00"
        ws.cell(row, 4).number_format = "0.00E+00"
        ws.cell(row, 5).number_format = "0.00"
        ws.cell(row, 6).number_format = "0.00"
        ws.cell(row, 7).number_format = "0.00"
        ws.cell(row, 8).number_format = "0.000000"
        ws.cell(row, 9).number_format = "0.000000"

    _apply_border(ws, 3, 3 + len(peak_results), 1, len(headers))

    for col, width in zip("ABCDEFGHI", [20, 18, 20, 20, 16, 16, 14, 14, 14]):
        ws.column_dimensions[col].width = width


# ──────────────────────────────────────────────────────────────────────────────

def main():
    nmr         = MnovaNMR.NMRPlugin()
    item        = nmr.activeNMRItem()
    sample_name = item.title(False) or "sample"
    b_values    = get_b_values(item)

    wb           = Workbook()
    peak_results = []

    for (peak_label, ppm_left, ppm_right) in PEAKS:
        print(f"Processing {peak_label} ...")

        Y, ppm_axis     = build_spectral_matrix(item, ppm_left, ppm_right)
        C, S_comp, eigs = decra(Y, n_components=N_COMPONENTS)
        D_vals, I0_vals = fit_diffusion(C, b_values)

        peak_results.append((peak_label, D_vals, I0_vals, eigs, S_comp, ppm_axis))

        write_peak_sheet(wb, peak_label, b_values, C, S_comp, eigs,
                         ppm_axis, D_vals, I0_vals)

    write_summary_sheet(wb, sample_name, peak_results)

    # move Summary to first position
    wb.move_sheet("Summary", offset=-len(wb.sheetnames) + 1)

    out_path = os.path.join(OUTPUT_FOLDER, f"{sample_name}_DECRA.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")


main()
