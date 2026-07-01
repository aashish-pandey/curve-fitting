import os
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.utils import get_column_letter
import MnovaNMR

# --- CONFIG ------------------------------------------------------------------

OUTPUT_FOLDER = r"C:\Users\you\Documents\NMR_Results"

# Same structure as before: (label, ppm_left, ppm_right, peak_type)
# peak_type: "sm" = pure small molecule
#            "poly" = pure polymer
#            "mixed" = mixed peak to separate
PEAKS = [
    ("sm_peak",    3.70, 3.60, "sm"),
    ("poly_peak",  1.50, 1.40, "poly"),
    ("mixed_peak", 1.30, 1.10, "mixed"),
]

N_COMPONENTS = 2   # polymer + small molecule

# MCR-ALS settings
MAX_ITER   = 500     # maximum iterations
TOL        = 1e-8    # convergence tolerance (change in fit between iterations)
N_RESTARTS = 3       # number of random restarts to avoid local minima

# -----------------------------------------------------------------------------


def get_b_values(nmr_item):
    """
    Return b-values (s/m^2) as a numpy array, one per gradient step.
    Replace with your own extraction logic.
    """
    spectra  = nmr_item.spectra()
    b_values = np.zeros(len(spectra))
    # --- YOUR CODE HERE ---
    return b_values


def integrate_peak(nmr_item, ppm_left, ppm_right):
    """
    For each gradient step, integrate the peak between ppm_left and ppm_right.
    Returns a 1D array of length n_gradients.
    """
    spectra  = nmr_item.spectra()
    ess      = spectra[0].coords[0]

    pt_left  = int(round(ess.ppmToPt(ppm_left)))
    pt_right = int(round(ess.ppmToPt(ppm_right)))
    pt_start = min(pt_left, pt_right)
    pt_end   = max(pt_left, pt_right)
    pts      = list(range(pt_start, pt_end + 1))

    integrals = np.zeros(len(spectra))
    for i, spc in enumerate(spectra):
        integrals[i] = sum(spc.reDataAt(pt) for pt in pts)

    return integrals


def build_Y(nmr_item, peaks):
    """
    Build Y matrix (n_gradients x n_peaks).
    Each column is the integrated intensity of one peak across all gradient steps.
    """
    columns = []
    for (label, ppm_left, ppm_right, peak_type) in peaks:
        col = integrate_peak(nmr_item, ppm_left, ppm_right)
        columns.append(col)
        print(f"  Integrated {label}: max={col.max():.2f}  min={col.min():.2f}")
    return np.column_stack(columns)   # (n_gradients, n_peaks)


# --- MCR-ALS -----------------------------------------------------------------

def _initialize_C(Y, peaks, n_components):
    """
    Initialize concentration profiles C using pure component columns.
    Pure SM column  → first  component (fast decay)
    Pure poly column → second component (slow decay)
    Falls back to SVD if no pure columns found.
    """
    M = Y.shape[0]
    C = np.zeros((M, n_components))

    sm_idx   = next((j for j, (_, _, _, t) in enumerate(peaks) if t == "sm"),   None)
    poly_idx = next((j for j, (_, _, _, t) in enumerate(peaks) if t == "poly"), None)

    if sm_idx is not None and poly_idx is not None:
        # use pure columns directly as initial concentration profiles
        # clip negatives to zero (non-negativity)
        C[:, 0] = np.clip(Y[:, sm_idx],   0, None)
        C[:, 1] = np.clip(Y[:, poly_idx], 0, None)
        # normalize each column to max=1 so scale is handled by S
        for k in range(n_components):
            mx = C[:, k].max()
            if mx > 0:
                C[:, k] /= mx
        print("  Initialization: using pure component columns (SM and polymer)")
    else:
        # fallback: SVD-based initialization
        U, s, Vt = np.linalg.svd(Y, full_matrices=False)
        C = np.abs(U[:, :n_components])
        print("  Initialization: SVD fallback (no pure columns found)")

    return C


def mcr_als(Y, peaks, n_components=2, max_iter=500, tol=1e-8, n_restarts=3):
    """
    MCR-ALS: Multivariate Curve Resolution - Alternating Least Squares.

    Decomposes Y (M x N) into C (M x K) and S (K x N) such that Y ~ C @ S
    with non-negativity constraints on both C and S.

    No assumption on b-value spacing — works with any x-axis.

    Parameters
    ----------
    Y           : (M, N)  rows = gradient steps, columns = peak integrals
    peaks       : list of (label, ppm_left, ppm_right, peak_type)
    n_components: K, number of components to extract
    max_iter    : maximum ALS iterations
    tol         : convergence threshold on relative change in fit error
    n_restarts  : number of restarts with different initializations

    Returns
    -------
    C           : (M, K)  concentration profiles (decay curves), non-negative
    S           : (K, N)  pure component spectra at b=0, non-negative
    residuals   : (M, N)  Y - C @ S
    fit_error   : final relative fit error
    n_iter      : iterations taken
    """
    Y = np.asarray(Y, dtype=float)
    M, N = Y.shape
    K    = n_components

    best_C, best_S, best_err = None, None, np.inf

    for restart in range(n_restarts):

        if restart == 0:
            C = _initialize_C(Y, peaks, K)
        else:
            # random restart with non-negative random matrix
            rng = np.random.default_rng(restart)
            C   = np.abs(rng.standard_normal((M, K)))
            for k in range(K):
                mx = C[:, k].max()
                if mx > 0:
                    C[:, k] /= mx

        prev_err = np.inf

        for iteration in range(max_iter):

            # --- Step 1: solve for S given C ---
            # Y = C @ S  =>  S = pinv(C) @ Y, then clip negatives
            S = np.linalg.lstsq(C, Y, rcond=None)[0]
            S = np.clip(S, 0, None)

            # --- Step 2: solve for C given S ---
            # Y = C @ S  =>  C = Y @ pinv(S), then clip negatives
            C = np.linalg.lstsq(S.T, Y.T, rcond=None)[0].T
            C = np.clip(C, 0, None)

            # --- convergence check ---
            err = np.linalg.norm(Y - C @ S) / (np.linalg.norm(Y) + 1e-12)
            if abs(prev_err - err) < tol:
                break
            prev_err = err

        if err < best_err:
            best_C, best_S, best_err = C.copy(), S.copy(), err
            best_iter = iteration + 1

        print(f"  Restart {restart+1}/{n_restarts}: err={err:.6f}  iters={iteration+1}")

    # sort components: largest initial value in C = slowest decay = polymer
    # (polymer persists longer so its profile stays higher across gradient steps)
    order = np.argsort(best_C[0, :])[::-1]
    best_C = best_C[:, order]
    best_S = best_S[order, :]

    residuals = Y - best_C @ best_S
    return best_C, best_S, residuals, best_err, best_iter


def fit_exponential(C, b_values):
    """
    Fit I(b) = I0 * exp(-D * b) to each component decay curve.
    Works with any b-value spacing.
    Returns D_values and I0_values.
    """
    b         = np.asarray(b_values, dtype=float)
    D_values  = np.zeros(C.shape[1])
    I0_values = np.zeros(C.shape[1])

    for k in range(C.shape[1]):
        decay = C[:, k]
        mask  = decay > 0
        if mask.sum() < 2:
            continue
        coeffs       = np.polyfit(b[mask], np.log(decay[mask]), deg=1)
        D_values[k]  = -coeffs[0]
        I0_values[k] =  np.exp(coeffs[1])

    return D_values, I0_values


# --- EXCEL HELPERS -----------------------------------------------------------

def _hdr(cell, value, bg="FF1F4E79"):
    cell.value     = value
    cell.font      = Font(bold=True, color="FFFFFFFF", name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg, end_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def _val(cell, value):
    cell.value = value
    cell.font  = Font(name="Arial", size=10)

def _thin_border():
    s = Side(style="thin", color="FF000000")
    return Border(left=s, right=s, top=s, bottom=s)

def _apply_border(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = _thin_border()


def write_results_sheet(wb, sample_name, peaks, b_values, C, S,
                        residuals, fit_error, n_iter, D_values, I0_values):
    ws            = wb.active
    ws.title      = "MCR-ALS Results"
    b             = np.asarray(b_values, dtype=float)
    comp_labels   = ["Small molecule", "Polymer"]
    fit_colors    = ["FFFF0000", "FF4472C4"]   # red, blue

    # --- header --------------------------------------------------------------
    ws.merge_cells("A1:H1")
    ws["A1"].value     = f"MCR-ALS Results - {sample_name}"
    ws["A1"].font      = Font(bold=True, size=13, name="Arial", color="FF1F4E79")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # --- component summary table (rows 3-6) ----------------------------------
    ws["A3"].value = "Component summary"
    ws["A3"].font  = Font(bold=True, name="Arial", size=10)

    for col, h in enumerate(["Component", "D (m2/s)", "I0", "Fit error", "Iterations", "Equation"], 1):
        _hdr(ws.cell(4, col), h)

    for k in range(N_COMPONENTS):
        r = 5 + k
        _val(ws.cell(r, 1), comp_labels[k])
        _val(ws.cell(r, 2), float(D_values[k]))
        _val(ws.cell(r, 3), float(I0_values[k]))
        _val(ws.cell(r, 4), float(fit_error))
        _val(ws.cell(r, 5), int(n_iter))
        _val(ws.cell(r, 6), f"I(b) = {I0_values[k]:.4f} * exp(-{D_values[k]:.4e} * b)")
        ws.cell(r, 2).number_format = "0.00E+00"
        ws.cell(r, 3).number_format = "0.0000"
        ws.cell(r, 4).number_format = "0.00E+00"

    _apply_border(ws, 4, 4 + N_COMPONENTS, 1, 6)

    # --- per-peak separation table (rows 9+) ---------------------------------
    ws["A9"].value = "Peak separation at b=0"
    ws["A9"].font  = Font(bold=True, name="Arial", size=10)

    sep_headers = ["Peak", "Type",
                   "SM contribution", "Polymer contribution",
                   "SM fraction (%)", "Polymer fraction (%)"]
    for col, h in enumerate(sep_headers, 1):
        _hdr(ws.cell(10, col), h)

    for j, (label, ppm_left, ppm_right, peak_type) in enumerate(peaks):
        r      = 11 + j
        sm_val = float(S[0, j])
        po_val = float(S[1, j])
        total  = sm_val + po_val
        _val(ws.cell(r, 1), label)
        _val(ws.cell(r, 2), peak_type)
        _val(ws.cell(r, 3), sm_val);  ws.cell(r, 3).number_format = "0.00"
        _val(ws.cell(r, 4), po_val);  ws.cell(r, 4).number_format = "0.00"
        sm_frac = (sm_val / total * 100) if total > 0 else 0.0
        po_frac = (po_val / total * 100) if total > 0 else 0.0
        _val(ws.cell(r, 5), sm_frac); ws.cell(r, 5).number_format = "0.00"
        _val(ws.cell(r, 6), po_frac); ws.cell(r, 6).number_format = "0.00"

    _apply_border(ws, 10, 10 + len(peaks), 1, len(sep_headers))

    # --- per-peak decay tables and charts ------------------------------------
    current_row   = 11 + len(peaks) + 3
    chart_anchor_col = 12

    for j, (peak_label, ppm_left, ppm_right, peak_type) in enumerate(peaks):

        ws.cell(current_row, 1).value = f"Decay curve: {peak_label} ({peak_type})"
        ws.cell(current_row, 1).font  = Font(bold=True, name="Arial", size=10)
        current_row += 1

        DR = current_row

        # headers: b | SM intensity | Poly intensity | SM fit | Poly fit
        _hdr(ws.cell(DR, 1), "b (s/m2)")
        _hdr(ws.cell(DR, 2), f"{comp_labels[0]} intensity", bg="FFc00000")
        _hdr(ws.cell(DR, 3), f"{comp_labels[1]} intensity", bg="FF1F4E79")
        _hdr(ws.cell(DR, 4), f"{comp_labels[0]} fit",       bg="FFc00000")
        _hdr(ws.cell(DR, 5), f"{comp_labels[1]} fit",       bg="FF1F4E79")

        for i, bv in enumerate(b):
            r = DR + 1 + i
            ws.cell(r, 1).value         = float(bv)
            ws.cell(r, 1).number_format = "0.00E+00"
            for k in range(N_COMPONENTS):
                # actual intensity = concentration profile * component spectrum value
                raw = float(C[i, k] * S[k, j])
                fit = float(I0_values[k] * np.exp(-D_values[k] * bv) * S[k, j])
                ws.cell(r, 2 + k).value         = raw
                ws.cell(r, 2 + k).number_format = "0.00"
                ws.cell(r, 4 + k).value         = fit
                ws.cell(r, 4 + k).number_format = "0.00"

        n_data_rows = DR + len(b)

        # chart
        chart              = ScatterChart()
        chart.scatterStyle = "smoothMarker"
        chart.title        = f"Decay: {peak_label} ({peak_type})"
        chart.x_axis.title = "b (s/m2)"
        chart.y_axis.title = "Intensity"
        chart.height       = 12
        chart.width        = 18
        chart.legend.position = "r"

        xvals = Reference(ws, min_col=1, min_row=DR + 1, max_row=n_data_rows)

        line_colors = ["FF0000", "4472C4"]
        for k in range(N_COMPONENTS):
            yvals  = Reference(ws, min_col=2 + k, min_row=DR + 1, max_row=n_data_rows)
            s_data = Series(yvals, xvals, title=f"{comp_labels[k]} data")
            s_data.marker.symbol = "circle"
            s_data.marker.size   = 5
            s_data.graphicalProperties.line.noFill = True
            chart.series.append(s_data)

            yfit  = Reference(ws, min_col=4 + k, min_row=DR + 1, max_row=n_data_rows)
            s_fit = Series(yfit, xvals, title=f"{comp_labels[k]} fit")
            s_fit.marker.symbol = "none"
            s_fit.graphicalProperties.line.solidFill = line_colors[k]
            s_fit.graphicalProperties.line.width     = 20000
            chart.series.append(s_fit)

        ws.add_chart(chart, f"{get_column_letter(chart_anchor_col)}{DR}")
        current_row = n_data_rows + 3

    # column widths
    for col, width in zip("ABCDEFGH", [22, 12, 20, 20, 18, 18, 44, 44]):
        ws.column_dimensions[col].width = width


# -----------------------------------------------------------------------------

def main():
    nmr         = MnovaNMR.NMRPlugin()
    item        = nmr.activeNMRItem()
    sample_name = item.title(False) or "sample"

    print(f"Sample: {sample_name}")
    print("Building Y matrix (one column per peak)...")

    b_values = get_b_values(item)
    Y        = build_Y(item, PEAKS)

    print(f"Y shape: {Y.shape}  ({Y.shape[0]} gradient steps x {Y.shape[1]} peaks)")
    print(f"b-values: {b_values}")

    print(f"\nRunning MCR-ALS (K={N_COMPONENTS}, max_iter={MAX_ITER}, restarts={N_RESTARTS})...")
    C, S, residuals, fit_error, n_iter = mcr_als(
        Y, PEAKS,
        n_components=N_COMPONENTS,
        max_iter=MAX_ITER,
        tol=TOL,
        n_restarts=N_RESTARTS,
    )

    D_values, I0_values = fit_exponential(C, b_values)

    print(f"\nFit error: {fit_error:.6f}  |  Iterations: {n_iter}")
    print("\nComponent summary:")
    comp_labels = ["Small molecule", "Polymer"]
    for k in range(N_COMPONENTS):
        print(f"  {comp_labels[k]}: D={D_values[k]:.4e} m2/s")

    print("\nPeak separation at b=0:")
    for j, (label, _, _, peak_type) in enumerate(PEAKS):
        total = S[0, j] + S[1, j]
        print(f"  {label} ({peak_type}):")
        for k in range(N_COMPONENTS):
            frac = S[k, j] / total * 100 if total > 0 else 0
            print(f"    {comp_labels[k]}: {S[k, j]:.4f}  ({frac:.1f}%)")

    wb = Workbook()
    write_results_sheet(wb, sample_name, PEAKS, b_values, C, S,
                        residuals, fit_error, n_iter, D_values, I0_values)

    out_path = os.path.join(OUTPUT_FOLDER, f"{sample_name}_MCRALS.xlsx")
    wb.save(out_path)
    print(f"\nSaved: {out_path}")


main()
