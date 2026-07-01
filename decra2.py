import os
import numpy as np
from numpy.linalg import eig, svd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import ScatterChart, Reference, Series
import MnovaNMR

# --- CONFIG ------------------------------------------------------------------

OUTPUT_FOLDER = r"C:\Users\you\Documents\NMR_Results"

# Define your peaks as (label, ppm_left, ppm_right, peak_type)
# peak_type: "sm" = pure small molecule
#            "poly" = pure polymer
#            "mixed" = mixed peak you want to separate
# Order does not matter, but you need at least one of each type.
PEAKS = [
    ("sm_peak",   3.70, 3.60, "sm"),
    ("poly_peak", 1.50, 1.40, "poly"),
    ("mixed_peak",1.30, 1.10, "mixed"),
]

N_COMPONENTS = 2   # polymer + small molecule

# -----------------------------------------------------------------------------


def get_b_values(nmr_item):
    """
    Return b-values (s/m^2) as a numpy array, one per gradient step.
    Replace with your own extraction logic.
    """
    spectra  = nmr_item.spectra()
    b_values = np.zeros(len(spectra))
    # --- YOUR CODE HERE ---
    return b_values


def integrate_peak(nmr_item, ppm_left, ppm_right):
    """
    For each gradient step, integrate the peak between ppm_left and ppm_right.
    Returns a 1D array of length n_gradients.
    Each value is the sum of real spectral intensities across the ppm window.
    """
    spectra  = nmr_item.spectra()
    ess      = spectra[0].coords[0]

    pt_left  = int(round(ess.ppmToPt(ppm_left)))
    pt_right = int(round(ess.ppmToPt(ppm_right)))
    pt_start = min(pt_left, pt_right)
    pt_end   = max(pt_left, pt_right)
    pts      = list(range(pt_start, pt_end + 1))

    integrals = np.zeros(len(spectra))
    for i, spc in enumerate(spectra):
        integrals[i] = sum(spc.reDataAt(pt) for pt in pts)

    return integrals


def build_Y(nmr_item, peaks):
    """
    Build Y matrix (n_gradients x n_peaks).
    Each column is the integrated intensity of one peak across all gradient steps.
    """
    columns = []
    for (label, ppm_left, ppm_right, peak_type) in peaks:
        col = integrate_peak(nmr_item, ppm_left, ppm_right)
        columns.append(col)
        print(f"  Integrated {label}: max={col.max():.2f}  min={col.min():.2f}")

    Y = np.column_stack(columns)   # (n_gradients, n_peaks)
    return Y


def decra(Y, n_components=2):
    """
    DECRA: Direct Exponential Curve Resolution Algorithm.

    Y : (M, N)  rows = gradient steps, columns = peaks (integrals).
                b-values must be uniformly spaced (constant delta_b between rows).

    Returns
    -------
    C           : (M, K)  decay profiles, normalised so C[0, :] = 1
    S_comp      : (K, N)  pure-component contributions at b=0, one value per peak
                           S_comp[0] = small molecule contribution to each peak
                           S_comp[1] = polymer contribution to each peak
    eigenvalues : (K,)    per-step decay factor, largest = slowest = polymer
    """
    Y  = np.asarray(Y, dtype=float)
    K  = n_components

    Y1 = Y[:-1, :]
    Y2 = Y[1:,  :]

    U, s, Vt = svd(Y1, full_matrices=False)
    U  = U[:,  :K]
    s  = s[    :K]
    Vt = Vt[:K, :]

    Phi_red      = U.T @ Y2 @ Vt.T @ np.diag(1.0 / s)
    eigenvalues, eigenvectors = eig(Phi_red)
    eigenvalues  = eigenvalues.real
    eigenvectors = eigenvectors.real

    order        = np.argsort(eigenvalues)[::-1]
    eigenvalues  = eigenvalues[order]
    eigenvectors = eigenvectors[:, order]

    C1 = U @ np.diag(s) @ eigenvectors
    C  = np.vstack([C1, C1[-1, :] * eigenvalues])
    C  = C / C[0, :]

    S_comp = np.linalg.pinv(C) @ Y   # (K, N) - one value per peak per component

    return C, S_comp, eigenvalues


def fit_diffusion(C, b_values):
    """
    Fit I(b) = I0 * exp(-D * b) to each component decay curve.
    Returns D_values and I0_values for each component.
    """
    b         = np.asarray(b_values, dtype=float)
    D_values  = np.zeros(C.shape[1])
    I0_values = np.zeros(C.shape[1])

    for k in range(C.shape[1]):
        decay = C[:, k]
        mask  = decay > 0
        if mask.sum() < 2:
            continue
        coeffs       = np.polyfit(b[mask], np.log(decay[mask]), deg=1)
        D_values[k]  = -coeffs[0]
        I0_values[k] =  np.exp(coeffs[1])

    return D_values, I0_values


# --- EXCEL HELPERS -----------------------------------------------------------

def _hdr(cell, value, bg="FF1F4E79"):
    cell.value     = value
    cell.font      = Font(bold=True, color="FFFFFFFF", name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg, end_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def _val(cell, value):
    cell.value = value
    cell.font  = Font(name="Arial", size=10)

def _thin_border():
    s = Side(style="thin", color="FF000000")
    return Border(left=s, right=s, top=s, bottom=s)

def _apply_border(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = _thin_border()


def write_results_sheet(wb, sample_name, peaks, b_values, C, S_comp,
                        eigenvalues, D_values, I0_values):
    """
    Write one sheet with:
      - Component summary (D, eigenvalue per component)
      - Per-peak separation table (what each component contributes to each peak)
      - Decay curve data + fit
      - Chart
    """
    ws     = wb.active
    ws.title = "DECRA Results"
    b      = np.asarray(b_values, dtype=float)
    comp_labels = ["Small molecule", "Polymer"]

    # header
    ws.merge_cells("A1:H1")
    ws["A1"].value     = f"DECRA Results - {sample_name}"
    ws["A1"].font      = Font(bold=True, size=13, name="Arial", color="FF1F4E79")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # component summary table (rows 3-5)
    ws["A3"].value = "Component summary"
    ws["A3"].font  = Font(bold=True, name="Arial", size=10)

    for col, h in enumerate(["Component", "D (m2/s)", "I0", "Eigenvalue", "Equation"], 1):
        _hdr(ws.cell(4, col), h)

    for k in range(N_COMPONENTS):
        r = 5 + k
        _val(ws.cell(r, 1), comp_labels[k])
        _val(ws.cell(r, 2), float(D_values[k]))
        _val(ws.cell(r, 3), float(I0_values[k]))
        _val(ws.cell(r, 4), float(eigenvalues[k]))
        _val(ws.cell(r, 5), f"I(b) = {I0_values[k]:.4f} * exp(-{D_values[k]:.4e} * b)")
        ws.cell(r, 2).number_format = "0.00E+00"
        ws.cell(r, 3).number_format = "0.0000"
        ws.cell(r, 4).number_format = "0.000000"

    _apply_border(ws, 4, 4 + N_COMPONENTS, 1, 5)

    # per-peak separation table (rows 9+)
    ws["A9"].value = "Peak separation at b=0"
    ws["A9"].font  = Font(bold=True, name="Arial", size=10)

    sep_headers = ["Peak", "Type", "Total integral"] + \
                  [f"{comp_labels[k]} contribution" for k in range(N_COMPONENTS)] + \
                  [f"{comp_labels[k]} fraction (%)" for k in range(N_COMPONENTS)]
    for col, h in enumerate(sep_headers, 1):
        _hdr(ws.cell(10, col), h)

    for j, (label, ppm_left, ppm_right, peak_type) in enumerate(peaks):
        r     = 11 + j
        total = sum(S_comp[k, j] for k in range(N_COMPONENTS))
        _val(ws.cell(r, 1), label)
        _val(ws.cell(r, 2), peak_type)
        _val(ws.cell(r, 3), float(total))
        ws.cell(r, 3).number_format = "0.00"
        for k in range(N_COMPONENTS):
            contrib = float(S_comp[k, j])
            _val(ws.cell(r, 4 + k), contrib)
            ws.cell(r, 4 + k).number_format = "0.00"
            frac = (contrib / total * 100) if total != 0 else 0.0
            _val(ws.cell(r, 4 + N_COMPONENTS + k), frac)
            ws.cell(r, 4 + N_COMPONENTS + k).number_format = "0.00"

    _apply_border(ws, 10, 10 + len(peaks), 1, len(sep_headers))

    # decay curve data table
    DR = 11 + len(peaks) + 3
    ws.cell(DR - 1, 1).value = "Decay curves"
    ws.cell(DR - 1, 1).font  = Font(bold=True, name="Arial", size=10)

    _hdr(ws.cell(DR, 1), "b (s/m2)")
    for k in range(N_COMPONENTS):
        _hdr(ws.cell(DR, 2 + k), f"{comp_labels[k]} C(b)")
        _hdr(ws.cell(DR, 2 + N_COMPONENTS + k), f"{comp_labels[k]} fit")

    for i, bv in enumerate(b):
        r = DR + 1 + i
        ws.cell(r, 1).value         = float(bv)
        ws.cell(r, 1).number_format = "0.00E+00"
        for k in range(N_COMPONENTS):
            ws.cell(r, 2 + k).value         = float(C[i, k])
            ws.cell(r, 2 + k).number_format = "0.0000"
            ws.cell(r, 2 + N_COMPONENTS + k).value         = float(I0_values[k] * np.exp(-D_values[k] * bv))
            ws.cell(r, 2 + N_COMPONENTS + k).number_format = "0.0000"

    n_rows_data = DR + len(b)

    # chart
    chart              = ScatterChart()
    chart.scatterStyle = "smoothMarker"
    chart.title        = "Decay curves - separated components"
    chart.x_axis.title = "b (s/m2)"
    chart.y_axis.title = "Normalised intensity"
    chart.height       = 14
    chart.width        = 22
    chart.legend.position = "r"

    fit_colors = ["FF0000", "4472C4"]

    for k in range(N_COMPONENTS):
        xvals  = Reference(ws, min_col=1,         min_row=DR + 1, max_row=n_rows_data)
        yvals  = Reference(ws, min_col=2 + k,     min_row=DR + 1, max_row=n_rows_data)
        s_data = Series(yvals, xvals, title=f"{comp_labels[k]} data")
        s_data.marker.symbol = "circle"
        s_data.marker.size   = 5
        s_data.graphicalProperties.line.noFill = True
        chart.series.append(s_data)

        yfit  = Reference(ws, min_col=2 + N_COMPONENTS + k, min_row=DR + 1, max_row=n_rows_data)
        s_fit = Series(yfit, xvals, title=f"{comp_labels[k]} fit")
        s_fit.marker.symbol = "none"
        s_fit.graphicalProperties.line.solidFill = fit_colors[k]
        s_fit.graphicalProperties.line.width     = 20000
        chart.series.append(s_fit)

    ws.add_chart(chart, f"A{n_rows_data + 3}")

    # column widths
    for col, width in zip("ABCDEFGH", [20, 12, 16, 22, 22, 18, 18, 44]):
        ws.column_dimensions[col].width = width


# -----------------------------------------------------------------------------

def main():
    nmr         = MnovaNMR.NMRPlugin()
    item        = nmr.activeNMRItem()
    sample_name = item.title(False) or "sample"

    print(f"Sample: {sample_name}")
    print("Building Y matrix (one column per peak)...")

    b_values = get_b_values(item)
    Y        = build_Y(item, PEAKS)

    print(f"Y matrix shape: {Y.shape}  ({Y.shape[0]} gradient steps x {Y.shape[1]} peaks)")

    # sanity check singular values
    from numpy.linalg import svd as _svd
    _, s, _ = _svd(Y, full_matrices=False)
    print(f"Singular values: {np.round(s[:4], 2)}")
    print(f"Ratio s1/s2: {s[0]/s[1]:.1f}  (lower = better separation)")

    print("Running DECRA...")
    C, S_comp, eigenvalues = decra(Y, n_components=N_COMPONENTS)
    D_values, I0_values    = fit_diffusion(C, b_values)

    print("\nResults:")
    comp_labels = ["Small molecule", "Polymer"]
    for k in range(N_COMPONENTS):
        print(f"  {comp_labels[k]}: D={D_values[k]:.4e} m2/s  eigenvalue={eigenvalues[k]:.6f}")

    print("\nPeak separation at b=0:")
    for j, (label, _, _, peak_type) in enumerate(PEAKS):
        print(f"  {label} ({peak_type}):")
        for k in range(N_COMPONENTS):
            print(f"    {comp_labels[k]}: {S_comp[k, j]:.4f}")

    wb = Workbook()
    write_results_sheet(wb, sample_name, PEAKS, b_values, C, S_comp,
                        eigenvalues, D_values, I0_values)

    out_path = os.path.join(OUTPUT_FOLDER, f"{sample_name}_DECRA.xlsx")
    wb.save(out_path)
    print(f"\nSaved: {out_path}")


main()
