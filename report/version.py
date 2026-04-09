#!/usr/bin/env python3
"""
NMR Curve Fitting Report Generator
Generates HTML report with SVG charts from Excel curve fitting data.
"""

import os
import openpyxl


def sanitize(text):
    """Replace special chars that cause encoding issues."""
    if not isinstance(text, str):
        return text
    replacements = {
        '·': '*',
        '−': '-',
        '²': '2',
        '³': '3',
        '±': '+/-',
        '×': 'x',
        '÷': '/',
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


def read_xlsx(filepath):
    """Read xlsx using openpyxl. Returns {sheet_name: [[row], ...]}"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        sheets[sheet_name] = data
    
    return sheets


def process_sheet(name, data):
    """Extract best model and plot data from sheet."""
    # Find summary section (row with '#' in first column)
    summary_idx = None
    for i, row in enumerate(data):
        if row and row[0] == '#':
            summary_idx = i
            break
    
    if summary_idx is None:
        print(f"    DEBUG: No summary section found")
        return None
    
    print(f"    DEBUG: Summary at row {summary_idx}")
    
    # Summary headers
    headers = [str(h).strip() if h else '' for h in data[summary_idx]]
    headers_lower = [h.lower() for h in headers]
    
    # Find column indices in summary
    def find_col(keywords):
        for i, h in enumerate(headers_lower):
            if any(k in h for k in keywords):
                return i
        return None
    
    idx_col = 0  # '#' is always first column
    name_col = find_col(['name'])
    formula_col = find_col(['formula'])
    r2_col = find_col(['r²', 'r2', 'r^2'])
    diff_col = find_col(['diffusion'])
    
    print(f"    DEBUG: Summary cols - name:{name_col} formula:{formula_col} r2:{r2_col} diff:{diff_col}")
    
    # Find best R² model
    best_r2, best_row = -1, None
    for row in data[summary_idx + 1:]:
        if not row or row[0] is None:
            break
        if r2_col is not None and len(row) > r2_col and row[r2_col] is not None:
            try:
                r2 = float(row[r2_col])
                if r2 > best_r2:
                    best_r2, best_row = r2, row
            except:
                pass
    
    if not best_row:
        print(f"    DEBUG: No valid R² found")
        return None
    
    model_name = str(best_row[name_col]) if name_col is not None else 'Unknown'
    model_num = int(best_row[idx_col]) if best_row[idx_col] is not None else 1
    print(f"    DEBUG: Best model #{model_num}: {model_name} R²={best_r2}")
    
    # Calculate fit/res column indices based on model #
    # #1 → cols 5,6 (F,G), #2 → cols 7,8 (H,I), etc.
    fit_col = 5 + (model_num - 1) * 2
    res_col = fit_col + 1
    
    # x and y are always first two columns (A=0, B=1)
    x_col = 0
    y_col = 1
    
    print(f"    DEBUG: x_col:{x_col} y_col:{y_col} fit_col:{fit_col} res_col:{res_col}")
    
    # Read data points
    x, y, fit, res = [], [], [], []
    for row in data[1:summary_idx]:
        if not row or row[0] is None:
            continue
        try:
            xv = float(row[x_col]) if row[x_col] is not None else None
            yv = float(row[y_col]) if row[y_col] is not None else None
            if xv is not None and yv is not None:
                x.append(xv)
                y.append(yv)
                fv = float(row[fit_col]) if len(row) > fit_col and row[fit_col] is not None else None
                rv = float(row[res_col]) if len(row) > res_col and row[res_col] is not None else None
                fit.append(fv)
                # Use residual from file, or calculate if None
                res.append(rv if rv is not None else (yv - fv if fv is not None else None))
        except:
            pass
    
    print(f"    DEBUG: {len(x)} data points")
    
    diff_coef = best_row[diff_col] if diff_col is not None and len(best_row) > diff_col else None
    formula = best_row[formula_col] if formula_col is not None and len(best_row) > formula_col else ''
    
    return {
        'name': name,
        'model': sanitize(str(model_name)),
        'formula': sanitize(str(formula)) if formula else '',
        'r2': best_r2,
        'diff_coef': diff_coef,
        'x': x, 'y': y, 'fit': fit, 'res': res
    }


def make_svg_chart(x, y, fit, res, title, r2, width=800, height=300):
    """Generate SVG chart with fit (left) and residual (right) side by side."""
    
    if not x or not y:
        return ""
    
    valid = [(xi, yi, fi, ri) for xi, yi, fi, ri in zip(x, y, fit, res) 
             if xi is not None and yi is not None]
    if not valid:
        return ""
    
    x_vals = [v[0] for v in valid]
    y_vals = [v[1] for v in valid]
    
    pad = 60
    chart_w = (width - 3 * pad) // 2
    chart_h = height - 2 * pad
    
    def scale(vals, size):
        mn, mx = min(vals), max(vals)
        rng = mx - mn if mx != mn else 1
        return mn, mx, lambda v: pad + (v - mn) / rng * size
    
    x_min, x_max, sx = scale(x_vals, chart_w)
    y_min, y_max, sy = scale(y_vals, chart_h)
    
    svg = [f'<svg width="{width}" height="{height}" xmlns="http://www.w3.org/2000/svg">']
    svg.append('<rect width="100%" height="100%" fill="white"/>')
    
    # Left chart: Fitted curve
    svg.append(f'<text x="{pad + chart_w//2}" y="20" text-anchor="middle" font-size="14" font-weight="bold">{title} (R² = {r2:.5f})</text>')
    
    svg.append(f'<line x1="{pad}" y1="{height-pad}" x2="{pad+chart_w}" y2="{height-pad}" stroke="black"/>')
    svg.append(f'<line x1="{pad}" y1="{pad}" x2="{pad}" y2="{height-pad}" stroke="black"/>')
    
    svg.append(f'<text x="{pad + chart_w//2}" y="{height-10}" text-anchor="middle" font-size="11">b-value</text>')
    svg.append(f'<text x="15" y="{height//2}" text-anchor="middle" font-size="11" transform="rotate(-90 15 {height//2})">Signal</text>')
    
    svg.append(f'<text x="{pad}" y="{height-pad+15}" text-anchor="middle" font-size="9">{x_min:.0f}</text>')
    svg.append(f'<text x="{pad+chart_w}" y="{height-pad+15}" text-anchor="middle" font-size="9">{x_max:.0f}</text>')
    svg.append(f'<text x="{pad-5}" y="{height-pad}" text-anchor="end" font-size="9">{y_min:.2f}</text>')
    svg.append(f'<text x="{pad-5}" y="{pad+5}" text-anchor="end" font-size="9">{y_max:.2f}</text>')
    
    for xi, yi in zip(x_vals, y_vals):
        px, py = sx(xi), height - pad - (sy(yi) - pad)
        svg.append(f'<circle cx="{px}" cy="{py}" r="4" fill="#4a90d9" opacity="0.7"/>')
    
    fit_points = [(xi, fi) for xi, fi in zip(x_vals, [v[2] for v in valid]) if fi is not None]
    fit_points.sort(key=lambda p: p[0])
    if fit_points:
        path = f'M {sx(fit_points[0][0])} {height - pad - (sy(fit_points[0][1]) - pad)}'
        for xi, fi in fit_points[1:]:
            path += f' L {sx(xi)} {height - pad - (sy(fi) - pad)}'
        svg.append(f'<path d="{path}" stroke="red" stroke-width="2" fill="none"/>')
    
    svg.append(f'<circle cx="{pad+chart_w-60}" cy="{pad+10}" r="4" fill="#4a90d9"/>')
    svg.append(f'<text x="{pad+chart_w-50}" y="{pad+14}" font-size="10">Data</text>')
    svg.append(f'<line x1="{pad+chart_w-65}" y1="{pad+25}" x2="{pad+chart_w-55}" y2="{pad+25}" stroke="red" stroke-width="2"/>')
    svg.append(f'<text x="{pad+chart_w-50}" y="{pad+29}" font-size="10">Fit</text>')
    
    # Right chart: Residuals
    offset = pad + chart_w + pad
    res_points = [(xi, ri) for xi, ri in zip(x_vals, [v[3] for v in valid]) if ri is not None]
    
    if res_points:
        r_vals = [p[1] for p in res_points]
        r_min, r_max, sr = scale(r_vals, chart_h)
        
        svg.append(f'<text x="{offset + chart_w//2}" y="20" text-anchor="middle" font-size="14" font-weight="bold">Residuals</text>')
        
        svg.append(f'<line x1="{offset}" y1="{height-pad}" x2="{offset+chart_w}" y2="{height-pad}" stroke="black"/>')
        svg.append(f'<line x1="{offset}" y1="{pad}" x2="{offset}" y2="{height-pad}" stroke="black"/>')
        
        zero_y = height - pad - (sr(0) - pad) if r_min <= 0 <= r_max else height // 2
        svg.append(f'<line x1="{offset}" y1="{zero_y}" x2="{offset+chart_w}" y2="{zero_y}" stroke="red" stroke-dasharray="5,3"/>')
        
        svg.append(f'<text x="{offset + chart_w//2}" y="{height-10}" text-anchor="middle" font-size="11">b-value</text>')
        svg.append(f'<text x="{offset-45}" y="{height//2}" text-anchor="middle" font-size="11" transform="rotate(-90 {offset-45} {height//2})">Residual</text>')
        
        svg.append(f'<text x="{offset}" y="{height-pad+15}" text-anchor="middle" font-size="9">{x_min:.0f}</text>')
        svg.append(f'<text x="{offset+chart_w}" y="{height-pad+15}" text-anchor="middle" font-size="9">{x_max:.0f}</text>')
        svg.append(f'<text x="{offset-5}" y="{height-pad}" text-anchor="end" font-size="9">{r_min:.3f}</text>')
        svg.append(f'<text x="{offset-5}" y="{pad+5}" text-anchor="end" font-size="9">{r_max:.3f}</text>')
        
        for xi, ri in res_points:
            px = offset + (sx(xi) - pad)
            py = height - pad - (sr(ri) - pad)
            svg.append(f'<circle cx="{px}" cy="{py}" r="4" fill="#2ecc71" opacity="0.7"/>')
    
    svg.append('</svg>')
    return '\n'.join(svg)


def generate_pdf_report(input_folder, output_folder=None):
    """Generate HTML report from all Excel files in a folder."""
    
    if output_folder is None:
        output_folder = input_folder
    
    if output_folder:
        os.makedirs(output_folder, exist_ok=True)
    
    output_filename = "curve_fitting_report.html"
    output_path = os.path.join(output_folder, output_filename) if output_folder else output_filename
    
    # Find all xlsx files
    xlsx_files = [f for f in os.listdir(input_folder) if f.endswith('.xlsx') and not f.startswith('~')]
    xlsx_files.sort()
    
    print(f"Found {len(xlsx_files)} Excel files in: {input_folder}")
    
    all_results = []
    
    for xlsx_file in xlsx_files:
        excel_path = os.path.join(input_folder, xlsx_file)
        sample_name = os.path.splitext(xlsx_file)[0]
        
        print(f"\nReading: {xlsx_file}")
        try:
            sheets = read_xlsx(excel_path)
            print(f"  Found {len(sheets)} sheets")
            
            for sheet_name, data in sheets.items():
                r = process_sheet(sheet_name, data)
                if r:
                    r['sample_name'] = sample_name
                    r['peak_name'] = sheet_name
                    all_results.append(r)
                    print(f"    {sheet_name}: {r['model']} (R²={r['r2']:.4f})")
                else:
                    all_results.append({'sample_name': sample_name, 'peak_name': sheet_name, 'error': True})
        except Exception as e:
            print(f"  Error: {e}")
    
    print(f"\nTotal results: {len(all_results)}")
    
    html = ['''<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<title>Curve Fitting Report</title>
<style>
body { font-family: Arial, sans-serif; margin: 40px; }
h1 { color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; }
h2 { color: #34495e; margin-top: 30px; }
h3 { color: #7f8c8d; }
table { border-collapse: collapse; width: 100%; margin: 20px 0; }
th { background: #3498db; color: white; padding: 10px; text-align: left; }
td { padding: 8px; border-bottom: 1px solid #ddd; }
tr:nth-child(even) { background: #f9f9f9; }
.chart-container { margin: 20px 0; page-break-inside: avoid; }
@media print { .chart-container { page-break-inside: avoid; } }
</style>
</head>
<body>
<h1>Curve Fitting Report</h1>
''']
    
    html.append('<h2>Section 1: Summary</h2>')
    html.append('<table>')
    html.append('<tr><th>Sample Name</th><th>Peak Name</th><th>Model (Equation)</th><th>R²</th><th>Diffusion Coefficient</th></tr>')
    
    for r in all_results:
        if r.get('error'):
            html.append(f'<tr><td>{r["sample_name"]}</td><td>{r["peak_name"]}</td><td colspan="3">Error processing</td></tr>')
        else:
            model_eq = f"{r['model']} ({r['formula']})" if r['formula'] else r['model']
            if len(model_eq) > 50:
                model_eq = model_eq[:47] + '...'
            diff = f"{r['diff_coef']:.4e}" if isinstance(r['diff_coef'], float) else 'N/A'
            html.append(f'<tr><td>{r["sample_name"]}</td><td>{r["peak_name"]}</td><td>{model_eq}</td>'
                       f'<td>{r["r2"]:.5f}</td><td>{diff}</td></tr>')
    
    html.append('</table>')
    
    html.append('<h2>Section 2: Fitted Curves & Residuals</h2>')
    
    for r in all_results:
        if r.get('error') or not r.get('x'):
            continue
        
        html.append(f'<div class="chart-container">')
        html.append(f'<h3>{r["sample_name"]} / {r["peak_name"]} - {r["model"]}</h3>')
        svg = make_svg_chart(r['x'], r['y'], r['fit'], r['res'], 'Fitted Curve', r['r2'])
        html.append(svg)
        html.append('</div>')
    
    html.append('</body></html>')
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(html))
    
    print(f"\nDone: {output_path}")
    print("Open in browser and print to PDF (Ctrl+P / Cmd+P)")
    return output_path


if __name__ == "__main__":
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    INPUT_FOLDER = SCRIPT_DIR  # folder containing xlsx files
    OUTPUT_FOLDER = SCRIPT_DIR  # where to save report
    
    generate_pdf_report(INPUT_FOLDER, OUTPUT_FOLDER)
