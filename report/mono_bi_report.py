#!/usr/bin/env python3
"""
NMR Curve Fitting Report Generator
Generates HTML report with SVG charts from Excel curve fitting data.

For each peak, shows BOTH Mono-Exponential and Bi-Exponential fits,
each with its own fit curve + residual chart, instead of picking a
single "best" model.
"""

import os
import openpyxl


def sanitize(text):
    """Replace special chars that cause encoding issues."""
    if not isinstance(text, str):
        return text
    replacements = {
        '·': '*',
        '−': '-',
        '²': '2',
        '³': '3',
        '±': '+/-',
        '×': 'x',
        '÷': '/',
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


def read_xlsx(filepath):
    """Read xlsx using openpyxl. Returns {sheet_name: [[row], ...]}"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        sheets[sheet_name] = data

    return sheets


def process_sheet(name, data):
    """Extract Mono-Exp and Bi-Exp model data + plot data from sheet."""
    # Find summary section (row with '#' in first column)
    summary_idx = None
    for i, row in enumerate(data):
        if row and row[0] == '#':
            summary_idx = i
            break

    if summary_idx is None:
        print(f"    DEBUG: No summary section found")
        return None

    print(f"    DEBUG: Summary at row {summary_idx}")

    # Summary headers
    headers = [str(h).strip() if h else '' for h in data[summary_idx]]
    headers_lower = [h.lower() for h in headers]

    # Find column indices in summary
    def find_col(keywords):
        for i, h in enumerate(headers_lower):
            if any(k in h for k in keywords):
                return i
        return None

    idx_col = 0  # '#' is always first column
    name_col = find_col(['name'])
    formula_col = find_col(['formula'])
    r2_col = find_col(['r²', 'r2', 'r^2'])
    diff_col = 6  # 7th column for diffusion coefficient

    print(f"    DEBUG: Summary cols - name:{name_col} formula:{formula_col} r2:{r2_col} diff:{diff_col}")

    # Find the Mono-Exp and Bi-Exp rows in the summary table
    mono_row, bi_row = None, None
    for row in data[summary_idx + 1:]:
        if not row or row[0] is None:
            break
        if name_col is None or len(row) <= name_col or row[name_col] is None:
            continue
        rname = str(row[name_col]).lower()
        if mono_row is None and 'mono' in rname:
            mono_row = row
        elif bi_row is None and 'bi' in rname:
            bi_row = row

    if mono_row is None and bi_row is None:
        print(f"    DEBUG: Neither Mono-Exp nor Bi-Exp found in summary")
        return None

    print(f"    DEBUG: mono_row found:{mono_row is not None} bi_row found:{bi_row is not None}")

    # Data headers (row 0)
    data_headers = [str(h).lower() if h else '' for h in data[0]]
    print(f"    DEBUG: Data headers: {data_headers[:10]}")

    # Find x (b_value) and y (I_norm) columns by header name
    x_col = None
    y_col = None
    for i, h in enumerate(data_headers):
        if 'b_value' in h or 'bvalue' in h:
            x_col = i
        if 'i_norm' in h or 'inorm' in h:
            y_col = i

    # Fixed column names for the two models we care about
    def find_data_col(colname):
        for i, h in enumerate(data_headers):
            if h == colname:
                return i
        return None

    mono_fit_col = find_data_col('monoexp_fit')
    mono_res_col = find_data_col('monoexp_res')
    bi_fit_col = find_data_col('biexp_fit')
    bi_res_col = find_data_col('biexp_res')

    print(f"    DEBUG: x_col:{x_col} y_col:{y_col} mono_fit:{mono_fit_col} mono_res:{mono_res_col} "
          f"bi_fit:{bi_fit_col} bi_res:{bi_res_col}")

    def get_val(row, col):
        if col is not None and len(row) > col and row[col] is not None:
            try:
                return float(row[col])
            except:
                return None
        return None

    # Read data points
    x, y = [], []
    mono_fit, mono_res = [], []
    bi_fit, bi_res = [], []

    for row in data[1:summary_idx]:
        if not row:
            continue
        # Skip truly empty rows (all None)
        if all(cell is None for cell in row[:5]):
            continue
        try:
            xv = float(row[x_col]) if x_col is not None and row[x_col] is not None else None
            yv = float(row[y_col]) if y_col is not None and row[y_col] is not None else None
            if xv is not None and yv is not None:
                x.append(xv)
                y.append(yv)

                mfv = get_val(row, mono_fit_col)
                mrv = get_val(row, mono_res_col)
                bfv = get_val(row, bi_fit_col)
                brv = get_val(row, bi_res_col)

                mono_fit.append(mfv)
                mono_res.append(mrv if mrv is not None else (yv - mfv if mfv is not None else None))
                bi_fit.append(bfv)
                bi_res.append(brv if brv is not None else (yv - bfv if bfv is not None else None))
        except:
            pass

    print(f"    DEBUG: {len(x)} data points")

    def build_model(row):
        if row is None:
            return None
        model_name = str(row[name_col]) if name_col is not None else 'Unknown'
        r2 = None
        if r2_col is not None and len(row) > r2_col and row[r2_col] is not None:
            try:
                r2 = float(row[r2_col])
            except:
                r2 = None
        diff_coef = row[diff_col] if diff_col is not None and len(row) > diff_col else None
        formula = row[formula_col] if formula_col is not None and len(row) > formula_col else ''
        try:
            diff_coef = float(diff_coef) if diff_coef is not None else None
        except:
            diff_coef = None
        return {
            'model': sanitize(str(model_name)),
            'formula': sanitize(str(formula)) if formula else '',
            'r2': r2,
            'diff_coef': diff_coef,
        }

    mono_info = build_model(mono_row)
    bi_info = build_model(bi_row)

    return {
        'name': name,
        'x': x, 'y': y,
        'mono': dict(mono_info, fit=mono_fit, res=mono_res) if mono_info else None,
        'bi': dict(bi_info, fit=bi_fit, res=bi_res) if bi_info else None,
    }


def make_svg_chart(x, y, fit, res, title, r2, width=800, height=300):
    """Generate SVG chart with fit (left) and residual (right) side by side."""

    if not x or not y:
        return ""

    valid = [(xi, yi, fi, ri) for xi, yi, fi, ri in zip(x, y, fit, res)
             if xi is not None and yi is not None]
    if not valid:
        return ""

    x_vals = [v[0] for v in valid]
    y_vals = [v[1] for v in valid]

    pad = 60
    chart_w = (width - 3 * pad) // 2
    chart_h = height - 2 * pad

    def scale(vals, size):
        mn, mx = min(vals), max(vals)
        rng = mx - mn if mx != mn else 1
        return mn, mx, lambda v: pad + (v - mn) / rng * size

    x_min, x_max, sx = scale(x_vals, chart_w)
    y_min, y_max, sy = scale(y_vals, chart_h)

    r2_label = f"{r2:.5f}" if isinstance(r2, float) else "N/A"

    svg = [f'<svg width="{width}" height="{height}" xmlns="http://www.w3.org/2000/svg">']
    svg.append('<rect width="100%" height="100%" fill="white"/>')

    # Left chart: Fitted curve
    svg.append(f'<text x="{pad + chart_w//2}" y="20" text-anchor="middle" font-size="14" font-weight="bold">{title} (R2 = {r2_label})</text>')

    svg.append(f'<line x1="{pad}" y1="{height-pad}" x2="{pad+chart_w}" y2="{height-pad}" stroke="black"/>')
    svg.append(f'<line x1="{pad}" y1="{pad}" x2="{pad}" y2="{height-pad}" stroke="black"/>')

    svg.append(f'<text x="{pad + chart_w//2}" y="{height-10}" text-anchor="middle" font-size="11">b-value</text>')
    svg.append(f'<text x="15" y="{height//2}" text-anchor="middle" font-size="11" transform="rotate(-90 15 {height//2})">Signal</text>')

    svg.append(f'<text x="{pad}" y="{height-pad+15}" text-anchor="middle" font-size="9">{x_min:.2f}</text>')
    svg.append(f'<text x="{pad+chart_w}" y="{height-pad+15}" text-anchor="middle" font-size="9">{x_max:.2f}</text>')
    svg.append(f'<text x="{pad-5}" y="{height-pad}" text-anchor="end" font-size="9">{y_min:.2f}</text>')
    svg.append(f'<text x="{pad-5}" y="{pad+5}" text-anchor="end" font-size="9">{y_max:.2f}</text>')

    for xi, yi in zip(x_vals, y_vals):
        px, py = sx(xi), height - pad - (sy(yi) - pad)
        svg.append(f'<circle cx="{px}" cy="{py}" r="4" fill="#4a90d9" opacity="0.7"/>')

    fit_points = [(xi, fi) for xi, fi in zip(x_vals, [v[2] for v in valid]) if fi is not None]
    fit_points.sort(key=lambda p: p[0])
    if fit_points:
        path = f'M {sx(fit_points[0][0])} {height - pad - (sy(fit_points[0][1]) - pad)}'
        for xi, fi in fit_points[1:]:
            path += f' L {sx(xi)} {height - pad - (sy(fi) - pad)}'
        svg.append(f'<path d="{path}" stroke="red" stroke-width="2" fill="none"/>')

    svg.append(f'<circle cx="{pad+chart_w-60}" cy="{pad+10}" r="4" fill="#4a90d9"/>')
    svg.append(f'<text x="{pad+chart_w-50}" y="{pad+14}" font-size="10">Data</text>')
    svg.append(f'<line x1="{pad+chart_w-65}" y1="{pad+25}" x2="{pad+chart_w-55}" y2="{pad+25}" stroke="red" stroke-width="2"/>')
    svg.append(f'<text x="{pad+chart_w-50}" y="{pad+29}" font-size="10">Fit</text>')

    # Right chart: Residuals
    offset = pad + chart_w + pad
    res_points = [(xi, ri) for xi, ri in zip(x_vals, [v[3] for v in valid]) if ri is not None]

    if res_points:
        r_vals = [p[1] for p in res_points]
        r_min, r_max, sr = scale(r_vals, chart_h)

        svg.append(f'<text x="{offset + chart_w//2}" y="20" text-anchor="middle" font-size="14" font-weight="bold">Residuals</text>')

        svg.append(f'<line x1="{offset}" y1="{height-pad}" x2="{offset+chart_w}" y2="{height-pad}" stroke="black"/>')
        svg.append(f'<line x1="{offset}" y1="{pad}" x2="{offset}" y2="{height-pad}" stroke="black"/>')

        zero_y = height - pad - (sr(0) - pad) if r_min <= 0 <= r_max else height // 2
        svg.append(f'<line x1="{offset}" y1="{zero_y}" x2="{offset+chart_w}" y2="{zero_y}" stroke="red" stroke-dasharray="5,3"/>')

        svg.append(f'<text x="{offset + chart_w//2}" y="{height-10}" text-anchor="middle" font-size="11">b-value</text>')
        svg.append(f'<text x="{offset-45}" y="{height//2}" text-anchor="middle" font-size="11" transform="rotate(-90 {offset-45} {height//2})">Residual</text>')

        svg.append(f'<text x="{offset}" y="{height-pad+15}" text-anchor="middle" font-size="9">{x_min:.2f}</text>')
        svg.append(f'<text x="{offset+chart_w}" y="{height-pad+15}" text-anchor="middle" font-size="9">{x_max:.2f}</text>')
        svg.append(f'<text x="{offset-5}" y="{height-pad}" text-anchor="end" font-size="9">{r_min:.4f}</text>')
        svg.append(f'<text x="{offset-5}" y="{pad+5}" text-anchor="end" font-size="9">{r_max:.4f}</text>')

        for xi, ri in res_points:
            px = offset + (sx(xi) - pad)
            py = height - pad - (sr(ri) - pad)
            svg.append(f'<circle cx="{px}" cy="{py}" r="4" fill="#2ecc71" opacity="0.7"/>')

    svg.append('</svg>')
    return '\n'.join(svg)


def generate_pdf_report(input_folder, output_folder=None):
    """Generate HTML report from all Excel files in a folder."""

    if output_folder is None:
        output_folder = input_folder

    if output_folder:
        os.makedirs(output_folder, exist_ok=True)

    output_filename = "curve_fitting_report.html"
    output_path = os.path.join(output_folder, output_filename) if output_folder else output_filename

    # Find all xlsx files
    xlsx_files = [f for f in os.listdir(input_folder) if f.endswith('.xlsx') and not f.startswith('~')]
    xlsx_files.sort()

    print(f"Found {len(xlsx_files)} Excel files in: {input_folder}")

    all_results = []

    for xlsx_file in xlsx_files:
        excel_path = os.path.join(input_folder, xlsx_file)
        sample_name = os.path.splitext(xlsx_file)[0]

        print(f"\nReading: {xlsx_file}")
        try:
            sheets = read_xlsx(excel_path)
            print(f"  Found {len(sheets)} sheets")

            for sheet_name, data in sheets.items():
                r = process_sheet(sheet_name, data)
                if r:
                    r['sample_name'] = sample_name
                    r['peak_name'] = sheet_name
                    all_results.append(r)
                    mono_r2 = r['mono']['r2'] if r['mono'] else None
                    bi_r2 = r['bi']['r2'] if r['bi'] else None
                    print(f"    {sheet_name}: Mono R2={mono_r2} Bi R2={bi_r2}")
                else:
                    all_results.append({'sample_name': sample_name, 'peak_name': sheet_name, 'error': True})
        except Exception as e:
            print(f"  Error: {e}")

    print(f"\nTotal results: {len(all_results)}")

    html = ['''<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<title>Curve Fitting Report</title>
<style>
body { font-family: Arial, sans-serif; margin: 40px; }
h1 { color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; }
h2 { color: #34495e; margin-top: 30px; }
h3 { color: #7f8c8d; }
table { border-collapse: collapse; width: 100%; margin: 20px 0; }
th { background: #3498db; color: white; padding: 10px; text-align: left; }
td { padding: 8px; border-bottom: 1px solid #ddd; }
tr:nth-child(even) { background: #f9f9f9; }
.chart-container { margin: 20px 0; page-break-inside: avoid; }
@media print { .chart-container { page-break-inside: avoid; } }
</style>
</head>
<body>
<h1>Curve Fitting Report</h1>
''']

    html.append('<h2>Section 1: Summary</h2>')
    html.append('<table>')
    html.append('<tr><th>Sample Name</th><th>Peak Name</th>'
                 '<th>Mono-Exp R2</th><th>Mono-Exp Diff. Coef.</th>'
                 '<th>Bi-Exp R2</th><th>Bi-Exp Diff. Coef.</th></tr>')

    for r in all_results:
        if r.get('error'):
            html.append(f'<tr><td>{r["sample_name"]}</td><td>{r["peak_name"]}</td><td colspan="4">Error processing</td></tr>')
            continue

        mono = r.get('mono')
        bi = r.get('bi')

        mono_r2 = f"{mono['r2']:.5f}" if mono and isinstance(mono['r2'], float) else 'N/A'
        mono_diff = f"{mono['diff_coef']:.4e}" if mono and isinstance(mono['diff_coef'], float) else 'N/A'
        bi_r2 = f"{bi['r2']:.5f}" if bi and isinstance(bi['r2'], float) else 'N/A'
        bi_diff = f"{bi['diff_coef']:.4e}" if bi and isinstance(bi['diff_coef'], float) else 'N/A'

        html.append(f'<tr><td>{r["sample_name"]}</td><td>{r["peak_name"]}</td>'
                     f'<td>{mono_r2}</td><td>{mono_diff}</td>'
                     f'<td>{bi_r2}</td><td>{bi_diff}</td></tr>')

    html.append('</table>')

    html.append('<h2>Section 2: Fitted Curves & Residuals</h2>')

    for r in all_results:
        if r.get('error') or not r.get('x'):
            continue

        # Mono-Exp chart pair
        if r.get('mono'):
            html.append(f'<div class="chart-container">')
            html.append(f'<h3>{r["sample_name"]} / {r["peak_name"]} - Mono-Exp</h3>')
            svg = make_svg_chart(r['x'], r['y'], r['mono']['fit'], r['mono']['res'],
                                  'Mono-Exp Fit', r['mono']['r2'])
            html.append(svg)
            html.append('</div>')

        # Bi-Exp chart pair
        if r.get('bi'):
            html.append(f'<div class="chart-container">')
            html.append(f'<h3>{r["sample_name"]} / {r["peak_name"]} - Bi-Exp</h3>')
            svg = make_svg_chart(r['x'], r['y'], r['bi']['fit'], r['bi']['res'],
                                  'Bi-Exp Fit', r['bi']['r2'])
            html.append(svg)
            html.append('</div>')

    html.append('</body></html>')

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(html))

    print(f"\nDone: {output_path}")
    print("Open in browser and print to PDF (Ctrl+P / Cmd+P)")
    return output_path


if __name__ == "__main__":
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    INPUT_FOLDER = SCRIPT_DIR
    OUTPUT_FOLDER = SCRIPT_DIR

    generate_pdf_report(INPUT_FOLDER, OUTPUT_FOLDER)
