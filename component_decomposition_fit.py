"""
dosy_curve_fit.py
─────────────────
Fits mono/bi-exponential decay curves to DOSY NMR decomposition Excel sheets.

Usage
-----
    from dosy_curve_fit import fit_dosy
    fit_dosy("output")   # scans output/ for *_component_Decomposition*.xlsx

For every sheet in every matching file the function:
  • reads  B_value (x),  I_norm_mixed, I_norm_sm, I_norm_poly (y's), in the
    ORIGINAL row order of the sheet (no sorting by B_value)
  • for each y-column independently: walking down the sheet in its existing
    row order, discards leading rows where the value is < 1; fitting starts
    from the first row >= 1
  • fits 5 curves:
      [1] I_norm_mixed → mono-exp → D_mixed_mono
      [2] I_norm_mixed → bi-exp   → D1_mixed_bi, D2_mixed_bi
      [3] I_norm_sm    → mono-exp → D_sm_mono
      [4] I_norm_poly  → mono-exp → D_poly_mono
      [5] I_norm_poly  → bi-exp   → D1_poly_bi, D2_poly_bi
  • params come from each column's trimmed (>=1) subset, but the fitted
    equation is evaluated ("extrapolated") back across ALL B_values / rows,
    so every fit column has a value for every row even though the fit only
    "saw" the trimmed subset during optimization
  • appends fitted-value columns and per-column trimmed raw-value columns
  • writes a results summary table below the data
  • embeds 5 individual ScatterCharts (one per fit: raw trimmed points +
    full-range fit line), then 4 overlay ScatterCharts (all Mixed-choice x
    SM-Mono x Poly-choice combinations = 2 x 1 x 2), each with 3 raw point
    series (different marker shape per column) + 3 fit lines
  • saves back to the same file

Dependencies: math, os, glob  (stdlib)  +  numpy  +  openpyxl
"""

import math, os, glob
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.axis import ChartLines


def fit_dosy(
    output_folder: str = "output",
    x_col:  str = "B_value",
    y_cols: tuple = ("I_norm_mixed", "I_norm_sm", "I_norm_poly"),
) -> None:

    y_mixed, y_sm, y_poly = y_cols

    # ── 1. Models ────────────────────────────────────────────────────────────

    def mono(x, A, D):
        return A * np.exp(np.clip(-D * x, -700, 700))

    def bi(x, A1, D1, A2, D2):
        return (A1 * np.exp(np.clip(-D1 * x, -700, 700))
              + A2 * np.exp(np.clip(-D2 * x, -700, 700)))

    # ── 2. Nelder-Mead (pure Python, no scipy) ───────────────────────────────

    def nelder_mead(f, p0, tol=1e-12, max_iter=20000):
        n = len(p0)
        s = [list(p0)]
        for i in range(n):
            v = list(p0); v[i] += 0.1 * abs(v[i]) if abs(v[i]) > 1e-10 else 1e-4
            s.append(v)
        sc = [f(v) for v in s]
        for _ in range(max_iter):
            o = sorted(range(n+1), key=lambda i: sc[i])
            s, sc = [s[i] for i in o], [sc[i] for i in o]
            if sc[-1] - sc[0] < tol: break
            c = [sum(s[i][j] for i in range(n)) / n for j in range(n)]
            xr = [c[j] + (c[j] - s[-1][j]) for j in range(n)]; fr = f(xr)
            if fr < sc[0]:
                xe = [c[j] + 2*(xr[j]-c[j]) for j in range(n)]; fe = f(xe)
                s[-1], sc[-1] = (xe,fe) if fe < fr else (xr,fr)
            elif fr < sc[-2]:
                s[-1], sc[-1] = xr, fr
            else:
                if fr < sc[-1]: s[-1], sc[-1] = xr, fr
                xc = [c[j] + 0.5*(s[-1][j]-c[j]) for j in range(n)]; fc_ = f(xc)
                if fc_ < sc[-1]: s[-1], sc[-1] = xc, fc_
                else:
                    for i in range(1, n+1):
                        s[i] = [s[0][j]+0.5*(s[i][j]-s[0][j]) for j in range(n)]
                        sc[i] = f(s[i])
        return np.array(s[0])

    def fit_curve(model, p0, bx, iy):
        """Run Nelder-Mead; return (params, fitted_y, R²) or None on failure."""
        if len(bx) < len(p0) + 1:
            return None
        def ssr(p):
            try:    return float(np.sum((iy - model(bx, *p))**2))
            except: return 1e30
        try:
            p = nelder_mead(ssr, p0)
            yp = model(bx, *p)
            ss_res = float(np.sum((iy - yp)**2))
            ss_tot = float(np.sum((iy - iy.mean())**2))
            r2 = 1 - ss_res/ss_tot if ss_tot > 0 else 0.0
            return p, yp, r2
        except:
            return None

    def warm_p0_mono(bx, iy):
        amp  = float(iy[np.argmin(bx)]) if len(iy) else 1.0
        amp  = amp if amp > 0 else 1.0
        span = float(bx.max() - bx.min()) if len(bx) > 1 else 1.0
        y0   = float(iy[np.argmin(bx)]); y0 = y0 if abs(y0) > 1e-30 else 1e-30
        yn   = float(iy[np.argmax(bx)]); yn = yn if abs(yn) > 1e-30 else 1e-30
        rate = abs(math.log(abs(yn/y0))) / span if span > 0 else 1e-9
        rate = rate if rate > 1e-30 else 1e-9
        return [amp, rate]

    def warm_p0_bi(bx, iy):
        amp, rate = warm_p0_mono(bx, iy)
        return [amp*0.6, rate, amp*0.4, rate*10]

    # ── 3. Excel style helpers ────────────────────────────────────────────────

    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")   # single header color, used everywhere
    W_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=9)   # header text (white, bold)
    N_FONT    = Font(name="Arial", size=9)                              # normal data text
    CTR       = Alignment(horizontal="center", vertical="center")

    def sc(cell, val, fill=None, font=None, align=CTR, nf=None):
        cell.value = val
        if fill: cell.fill = fill
        cell.font = font or N_FONT
        cell.alignment = align
        if nf: cell.number_format = nf

    # ── 4. Chart builders ────────────────────────────────────────────────────

    def make_scatter(ws, title, xref, yref_data, yref_fit, marker_color, marker_symbol, anchor):
        ch = ScatterChart(); ch.scatterStyle = "lineMarker"
        ch.title = title; ch.width = 14; ch.height = 10
        ch.x_axis.title = x_col; ch.y_axis.title = "I_norm"
        ch.x_axis.numFmt = "0.00E+00"; ch.style = 2
        ch.x_axis.delete = False; ch.y_axis.delete = False
        ch.x_axis.majorTickMark = "out"; ch.y_axis.majorTickMark = "out"
        ch.x_axis.majorGridlines = ChartLines()

        raw = Series(yref_data, xref, title="Data")
        raw.marker.symbol = marker_symbol; raw.marker.size = 5
        raw.marker.graphicalProperties.solidFill = marker_color
        raw.graphicalProperties.line.noFill = True
        ch.series.append(raw)

        fit = Series(yref_fit, xref, title="Fit")
        fit.marker.symbol = "none"
        fit.graphicalProperties.line.solidFill = "FF0000"
        fit.graphicalProperties.line.width = 18000
        ch.series.append(fit)

        ws.add_chart(ch, anchor)

    # ── 4b. Overlay chart builder ────────────────────────────────────────────

    def make_overlay(ws, title, xref, curve_defs, raw_defs, anchor):
        """
        curve_defs: list of (col_idx, label, hex_color) — one LINE per fit.
        raw_defs:   list of (col_idx, label, hex_color, marker_symbol) —
                    one MARKER-only series per raw column (blanks skipped).
        """
        ch = ScatterChart(); ch.scatterStyle = "lineMarker"
        ch.title = title; ch.width = 16; ch.height = 10
        ch.x_axis.title = x_col; ch.y_axis.title = "I_norm"
        ch.x_axis.numFmt = "0.00E+00"; ch.style = 2
        ch.x_axis.delete = False; ch.y_axis.delete = False
        ch.x_axis.majorTickMark = "out"; ch.y_axis.majorTickMark = "out"
        ch.x_axis.majorGridlines = ChartLines()

        for col_idx, label, color, symbol in raw_defs:
            yref = Reference(ws, min_col=col_idx, min_row=xref.min_row, max_row=xref.max_row)
            s = Series(yref, xref, title=label)
            s.marker.symbol = symbol; s.marker.size = 6
            s.marker.graphicalProperties.solidFill = color
            s.graphicalProperties.line.noFill = True
            ch.series.append(s)

        for col_idx, label, color in curve_defs:
            yref = Reference(ws, min_col=col_idx, min_row=xref.min_row, max_row=xref.max_row)
            s = Series(yref, xref, title=label)
            s.marker.symbol = "none"
            s.graphicalProperties.line.solidFill = color
            s.graphicalProperties.line.width = 20000
            ch.series.append(s)

        ws.add_chart(ch, anchor)

    # ── 5. Per-sheet processing ───────────────────────────────────────────────

    def process_sheet(ws, fname):
        ws._charts = []

        # 5a. locate header row ───────────────────────────────────────────────
        xl = x_col.strip().lower()
        yls = [c.strip().lower() for c in (y_mixed, y_sm, y_poly)]
        hrow = xi = None
        yi_map = {}
        for row in ws.iter_rows(min_row=1, max_row=min(20, ws.max_row)):
            found_x = None
            found_y = {}
            for cell in row:
                if cell.value is None: continue
                raw = cell.value
                if isinstance(raw, (list, tuple)): raw = raw[0] if raw else None
                if raw is None: continue
                label = str(raw).strip().lower()
                if label == xl: found_x = cell.column
                elif label in yls: found_y[label] = cell.column
            if found_x is not None and len(found_y) == 3:
                hrow = row[0].row
                xi = found_x
                yi_map = found_y
                break
        if hrow is None:
            print(f"  [SKIP] {ws.title}: headers not found"); return

        yi_mixed = yi_map[y_mixed.strip().lower()]
        yi_sm    = yi_map[y_sm.strip().lower()]
        yi_poly  = yi_map[y_poly.strip().lower()]

        # color the whole existing header row (not just the columns we add later)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=hrow, column=col)
            if cell.value is not None:
                cell.fill = HDR_FILL
                cell.font = W_FONT
                cell.alignment = CTR

        # 5b. read data ────────────────────────────────────────────────────────
        row_nums, b_raw, mixed_raw, sm_raw, poly_raw = [], [], [], [], []
        for row in ws.iter_rows(min_row=hrow+1, max_row=ws.max_row):
            bv = row[xi-1].value
            mv = row[yi_mixed-1].value
            sv = row[yi_sm-1].value
            pv = row[yi_poly-1].value
            if bv is None: continue
            row_nums.append(row[0].row)
            b_raw.append(float(bv))
            mixed_raw.append(float(mv) if mv is not None else np.nan)
            sm_raw.append(float(sv) if sv is not None else np.nan)
            poly_raw.append(float(pv) if pv is not None else np.nan)

        row_nums = np.array(row_nums)
        b_raw = np.array(b_raw)

        def trim(y_arr):
            """Walk rows in original sheet order; discard leading rows where y < 1."""
            start = int(np.argmax(y_arr >= 1))
            return row_nums[start:], b_raw[start:], y_arr[start:]

        rn_mixed, b_mixed, i_mixed = trim(np.array(mixed_raw))
        rn_sm,    b_sm,    i_sm    = trim(np.array(sm_raw))
        rn_poly,  b_poly,  i_poly  = trim(np.array(poly_raw))

        # 5c. run fits ─────────────────────────────────────────────────────────
        fits = {
            "mixed_mono": fit_curve(mono, warm_p0_mono(b_mixed, i_mixed), b_mixed, i_mixed),
            "mixed_bi":   fit_curve(bi,   warm_p0_bi(b_mixed, i_mixed),   b_mixed, i_mixed),
            "sm_mono":    fit_curve(mono, warm_p0_mono(b_sm, i_sm),       b_sm, i_sm),
            "poly_mono":  fit_curve(mono, warm_p0_mono(b_poly, i_poly),   b_poly, i_poly),
            "poly_bi":    fit_curve(bi,   warm_p0_bi(b_poly, i_poly),     b_poly, i_poly),
        }

        # 5d. write fitted columns + trimmed raw columns ────────────────────────
        # Params are fit on each column's trimmed (>=1) subset, but the fitted
        # equation is then evaluated over ALL B_values / ALL rows so the curve
        # spans the full plot range, not just the trimmed window.
        FIT_HDRS   = ["Fit_Mixed_Mono","Fit_Mixed_Bi","Fit_SM_Mono","Fit_Poly_Mono","Fit_Poly_Bi"]
        FIT_KEYS   = ["mixed_mono","mixed_bi","sm_mono","poly_mono","poly_bi"]
        FIT_MODELS = [mono, bi, mono, mono, bi]
        FIT_ROWS   = [row_nums, row_nums, row_nums, row_nums, row_nums]
        FIT_BX     = [b_raw, b_raw, b_raw, b_raw, b_raw]
        # Place output columns safely past every existing column in the sheet
        # (not a hardcoded offset) so we never overwrite B_value or any other
        # pre-existing data — this was the bug that clobbered B_value before.
        COL_OFF = max(ws.max_column, xi, yi_mixed, yi_sm, yi_poly) + 2

        key_to_col = {}
        for k, (key, hdr, model, rn_k, bx_k) in enumerate(zip(FIT_KEYS, FIT_HDRS, FIT_MODELS, FIT_ROWS, FIT_BX)):
            col = COL_OFF + k
            key_to_col[key] = col
            hcell = ws.cell(row=hrow, column=col)
            sc(hcell, hdr, fill=HDR_FILL, font=W_FONT)
            ws.column_dimensions[get_column_letter(col)].width = 16
            if fits[key] is None: continue
            params = fits[key][0]
            yp = model(bx_k, *params)
            for r_, v_ in zip(rn_k, yp):
                c = ws.cell(row=int(r_), column=col)
                c.value = round(float(v_), 8)
                c.number_format = "0.00000000"
                c.font = N_FONT; c.alignment = CTR

        RAW_COLS = {"mixed": (COL_OFF+5, rn_mixed, i_mixed, "Raw_Mixed"),
                    "sm":    (COL_OFF+6, rn_sm,    i_sm,    "Raw_SM"),
                    "poly":  (COL_OFF+7, rn_poly,  i_poly,  "Raw_Poly")}
        for name, (col, rn_k, iy_k, hdr) in RAW_COLS.items():
            sc(ws.cell(row=hrow, column=col), hdr, fill=HDR_FILL, font=W_FONT)
            ws.column_dimensions[get_column_letter(col)].width = 14
            for r_, v_ in zip(rn_k, iy_k):
                c = ws.cell(row=int(r_), column=col)
                c.value = round(float(v_), 8)
                c.number_format = "0.00000000"
                c.font = N_FONT; c.alignment = CTR

        # 5e. summary table ────────────────────────────────────────────────────
        last_row = int(row_nums.max())
        sr = last_row + 3

        ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=6)
        sc(ws.cell(row=sr, column=1),
           f"Curve Fit Results — {ws.title}", fill=HDR_FILL, font=W_FONT)
        sr += 1

        for ci_, h in enumerate(["Fit","Equation","D Coeff(s) [m²/s]","A","R²","Notes"], 1):
            sc(ws.cell(row=sr, column=ci_), h, fill=HDR_FILL, font=W_FONT)
        sr += 1

        def d_val(key, idx):
            if fits[key] is None: return "failed"
            return float(fits[key][0][idx])

        def r2_val(key):
            if fits[key] is None: return "—"
            return round(fits[key][2], 6)

        EQ_MONO = "I = A · exp(−D·B)"
        EQ_BI   = "I = A1·exp(−D1·B) + A2·exp(−D2·B)"

        summary_rows = [
            ("Mixed → Mono",    EQ_MONO, d_val("mixed_mono",1), d_val("mixed_mono",0), r2_val("mixed_mono"), ""),
            ("Mixed → Bi (D1)", EQ_BI,   d_val("mixed_bi",1),   d_val("mixed_bi",0),   r2_val("mixed_bi"),  "fast component"),
            ("Mixed → Bi (D2)", EQ_BI,   d_val("mixed_bi",3),   d_val("mixed_bi",2),   r2_val("mixed_bi"),  "slow component"),
            ("SM → Mono",       EQ_MONO, d_val("sm_mono",1),    d_val("sm_mono",0),    r2_val("sm_mono"),   ""),
            ("Poly → Mono",     EQ_MONO, d_val("poly_mono",1),  d_val("poly_mono",0),  r2_val("poly_mono"), ""),
            ("Poly → Bi (D1)",  EQ_BI,   d_val("poly_bi",1),    d_val("poly_bi",0),    r2_val("poly_bi"),   "fast component"),
            ("Poly → Bi (D2)",  EQ_BI,   d_val("poly_bi",3),    d_val("poly_bi",2),    r2_val("poly_bi"),   "slow component"),
        ]

        for row_data in summary_rows:
            for ci_, val in enumerate(row_data, 1):
                cell = ws.cell(row=sr, column=ci_)
                if isinstance(val, float):
                    sc(cell, val, fill=None, nf="0.00000E+00")
                else:
                    sc(cell, val, fill=None)
            sr += 1

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 18

        # 5f. overlay charts — 4 combos: Mixed(Mono/Bi) x SM(Mono) x Poly(Mono/Bi) ──
        dmin, dmax = int(row_nums.min()), int(row_nums.max())
        xref_all = Reference(ws, min_col=xi, min_row=dmin, max_row=dmax)

        KEY_LABEL = {"mixed_mono": "Mixed-Mono", "mixed_bi": "Mixed-Bi",
                     "sm_mono": "SM-Mono", "poly_mono": "Poly-Mono", "poly_bi": "Poly-Bi"}
        KEY_COLOR = {"mixed_mono": "4472C4", "mixed_bi": "1F4E79",
                     "sm_mono": "70AD47", "poly_mono": "ED7D31", "poly_bi": "C00000"}

        raw_defs = [
            (RAW_COLS["mixed"][0], "Mixed-Raw", "4472C4", "circle"),
            (RAW_COLS["sm"][0],    "SM-Raw",    "70AD47", "triangle"),
            (RAW_COLS["poly"][0],  "Poly-Raw",  "ED7D31", "diamond"),
        ]

        overlay_combos = [
            ("Mixed-Mono + SM-Mono + Poly-Mono", ["mixed_mono", "sm_mono", "poly_mono"]),
            ("Mixed-Bi + SM-Mono + Poly-Mono",   ["mixed_bi",   "sm_mono", "poly_mono"]),
            ("Mixed-Mono + SM-Mono + Poly-Bi",   ["mixed_mono", "sm_mono", "poly_bi"]),
            ("Mixed-Bi + SM-Mono + Poly-Bi",     ["mixed_bi",   "sm_mono", "poly_bi"]),
        ]

        CHART_ROW_SPACING = 26
        overlay_anchor_row = sr + 2

        # 5f-1. overlay charts (multi-curve) first — 4 combos, stacked in column A
        placed = 0
        for title, keys in overlay_combos:
            if any(fits[k_] is None for k_ in keys):
                print(f"    [SKIP OVERLAY] {title}: missing fit(s)")
                continue
            curve_defs = [(key_to_col[k_], KEY_LABEL[k_], KEY_COLOR[k_]) for k_ in keys]
            anchor = f"A{overlay_anchor_row + placed*CHART_ROW_SPACING}"
            make_overlay(ws, title, xref_all, curve_defs, raw_defs, anchor)
            placed += 1

        # 5f-2. individual charts second — one per fit, 3 rows x 2 columns grid
        individual_configs = [
            ("Mixed — Mono-Exp", "mixed_mono", RAW_COLS["mixed"][0], "4472C4", "circle"),
            ("Mixed — Bi-Exp",   "mixed_bi",   RAW_COLS["mixed"][0], "4472C4", "circle"),
            ("SM — Mono-Exp",    "sm_mono",    RAW_COLS["sm"][0],    "70AD47", "triangle"),
            ("Poly — Mono-Exp",  "poly_mono",  RAW_COLS["poly"][0],  "ED7D31", "diamond"),
            ("Poly — Bi-Exp",    "poly_bi",    RAW_COLS["poly"][0],  "ED7D31", "diamond"),
        ]
        individual_anchor_row = overlay_anchor_row + placed*CHART_ROW_SPACING
        GRID_COLS = ["A", "J"]   # 2 columns per row of the grid
        placed_ind = 0
        for title, key, raw_col, color, symbol in individual_configs:
            if fits[key] is None:
                print(f"    [SKIP CHART] {title}: fit failed")
                continue
            yref_data = Reference(ws, min_col=raw_col, min_row=dmin, max_row=dmax)
            yref_fit  = Reference(ws, min_col=key_to_col[key], min_row=dmin, max_row=dmax)
            grid_row, grid_col = divmod(placed_ind, 2)
            anchor = f"{GRID_COLS[grid_col]}{individual_anchor_row + grid_row*CHART_ROW_SPACING}"
            make_scatter(ws, title, xref_all, yref_data, yref_fit, color, symbol, anchor)
            placed_ind += 1

        print(f"  [OK] {ws.title} | N={len(b_raw)} mixed_kept={len(b_mixed)} "
              f"sm_kept={len(b_sm)} poly_kept={len(b_poly)} | "
              f"individual={placed_ind} overlays={placed}")

    # ── 6. File discovery & main loop ─────────────────────────────────────────

    pattern  = os.path.join(output_folder, "**", "*_component_Decomposition*.xlsx")
    files    = list(set(
        glob.glob(pattern, recursive=True)
        + glob.glob(os.path.join(output_folder, "*_component_Decomposition*.xlsx"))
    ))

    if not files:
        print(f"No matching files found in '{output_folder}'"); return

    print(f"Found {len(files)} file(s).")
    for fp_ in sorted(files):
        print(f"\n[FILE] {os.path.basename(fp_)}")
        wb = openpyxl.load_workbook(fp_)
        for sname in wb.sheetnames:
            process_sheet(wb[sname], os.path.basename(fp_))
        wb.save(fp_)
        print(f"  Saved → {fp_}")


if __name__ == "__main__":
    import sys
    fit_dosy(sys.argv[1] if len(sys.argv) > 1 else "output")
